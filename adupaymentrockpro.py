# adupaymentrock.py (Final Version - All Steps Included)
from database_manager import DatabaseManager
import re
import os
import json
import io
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Tuple
import uuid
import logging
import asyncio
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
import sys
import subprocess
import shutil
import datetime as dt  # Renamed for clarity
import base64  # Needed for embedding charts in PDF

# ----------------------- Config & Logging -----------------------
TELEGRAM_BOT_TOKEN = os.getenv('TELEGRAM_BOT_TOKEN', '')
# --- (!!!) NEW: Persistent Disk for Bot State File (!!!) ---
# Database က PostgreSQL မှာမို့၊ ဒီ Disk ကို 'bot_persistence' file အတွက်ပဲ သုံးပါမယ်။
DATA_ROOT = os.getenv('DATA_DIR', '.') # Render က /app/data ကို ဖတ်ပါမယ်
PERSISTENCE_FILE_PATH = os.path.join(DATA_ROOT, 'bot_persistence')
EXPORT_DIR = os.path.join(DATA_ROOT, 'exports') # Export file တွေပါ ဒီထဲပဲ ထားပါ

print(f"--- 💡 BOT STATE STORAGE INITIALIZED ---")
print(f"Using DATA_ROOT (for persistence): {DATA_ROOT}")
print(f"Persistence File Path: {PERSISTENCE_FILE_PATH}")
# --- (!!!) End of New Config (!!!) ---

# --- NEW: DatabaseManager ---
# database_manager.py file ကို ခေါ်တဲ့ နေရာတွေမှာ DB_PATH ကို သုံးရပါမယ်။
from database_manager import DatabaseManager

# --- NEW FONT INSTALLER (Python Method) ---

# --- NEW: DatabaseManager ---
# -----------------------------------------------
# 💡 NEW FONT INSTALLER (Python Method)
# -----------------------------------------------
def install_myanmar_font():
    """
    Render.com server ပေါ်တွင် မြန်မာ Font ကို Python code ဖြင့်
    တိုက်ရိုက် သွင်းပေးသော function။
    """
    print("--- 💡 Starting Font Installation Check ---")

    try:
        # 1. Project folder ထဲက font file လမ်းကြောင်း
        current_dir = os.path.dirname(os.path.abspath(__file__))
        project_font_path = os.path.join(current_dir, 'fonts', 'Pyidaungsu-Regular.ttf')

        # 2. Server မှာ Font သွင်းမယ့် နေရာ (~/.fonts)
        font_dir = os.path.expanduser("~/.fonts")
        target_font_path = os.path.join(font_dir, "Pyidaungsu-Regular.ttf")

        # 3. Server မှာ Font ရှိ၊ မရှိ စစ်ဆေးခြင်း
        if os.path.exists(target_font_path):
            print(f"✅ Font '{target_font_path}' already exists. Skipping installation.")
            return True

        print(f"ℹ️ Font not found. Starting installation...")

        # 4. Project folder ထဲမှာ Font file ရှိ၊ မရှိ စစ်ဆေးခြင်း
        if not os.path.exists(project_font_path):
            print(f"❌ CRITICAL: Source font '{project_font_path}' not found. Cannot install font.")
            return False

        # 5. ~/.fonts directory ကို ဆောက်ခြင်း
        print(f"Creating directory '{font_dir}'...")
        os.makedirs(font_dir, exist_ok=True)

        # 6. Font file ကို copy ကူးထည့်ခြင်း
        print(f"Copying '{project_font_path}' to '{font_dir}'...")
        shutil.copy(project_font_path, font_dir)

        # 7. Font cache ကို refresh လုပ်ခြင်း
        print("Refreshing font cache using 'fc-cache'...")
        subprocess.run(["fc-cache", "-fv"], check=True)

        print("✅--- Font Installation Succeeded ---✅")
        return True

    except Exception as e:
        print(f"❌--- Font Installation Failed ---❌")
        print(f"Error: {e}")
        return False

# -----------------------------------------------
# အခုပဲ Font သွင်းတဲ့ function ကို Run ပါ
install_myanmar_font()
# -----------------------------------------------


# --- The rest of your code ... ---
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s - %(lineno)d', level=logging.INFO)
logger = logging.getLogger(__name__)

# --- Third-party Libs ---
try:
    from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton, CallbackQuery
    from telegram.ext import Application, CommandHandler, MessageHandler, filters, CallbackQueryHandler, ContextTypes, PicklePersistence
    from telegram.constants import ParseMode
    TELEGRAM_AVAILABLE = True
except ImportError:
    TELEGRAM_AVAILABLE = False
    logger.critical(
        "❌ python-telegram-bot library not found. Please run 'pip install python-telegram-bot pandas plotly kaleido weasyprint openpyxl sqlalchemy'")
    sys.exit(1)

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    logger.critical("❌ pandas library not found.")
    sys.exit(1)

# Excel libs
try:
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# PDF libs: WeasyPrint
try:
    from weasyprint import HTML, CSS
    WEASYPRINT_AVAILABLE = True
except (ImportError, OSError) as e:
    logger.critical(
        f"WeasyPrint is not available, PDF/Chart export will fail. Error: {e}")
    WEASYPRINT_AVAILABLE = False

# Chart Lib: Plotly
try:
    import plotly.graph_objects as go
    import plotly.io as pio
    import plotly.colors as px_colors
    try:
        import kaleido
        KALEIDO_AVAILABLE = True
    except ImportError:
        KALEIDO_AVAILABLE = False
        logger.warning(
            "Kaleido not found ('pip install kaleido'). Plotly image export might be slower or fail.")

    PLOTLY_AVAILABLE = True
    pio.templates.default = "plotly_white"
except ImportError:
    PLOTLY_AVAILABLE = False
    logger.warning(
        "Plotly not found ('pip install plotly kaleido'). Chart generation disabled.")

# --- NEW: SQLAlchemy Check ---
try:
    import sqlalchemy
    SQLALCHEMY_AVAILABLE = True
except ImportError:
    SQLALCHEMY_AVAILABLE = False
    logger.critical(
        "❌ sqlalchemy library not found. Please run 'pip install sqlalchemy'")
    sys.exit(1)


# --- Font Path Setup ---
CURRENT_DIR = os.path.dirname(os.path.abspath(sys.argv[0])) if os.path.dirname(
    os.path.abspath(sys.argv[0])) else os.getcwd()
FONT_PATH = os.path.join(CURRENT_DIR, 'fonts', 'Pyidaungsu-Regular.ttf')
CLEAN_FONT_PATH = FONT_PATH.replace(os.path.sep, '/')

# ----------------------- Texts (Myanmar/English) -----------------------
TEXTS = {
    "start_welcome": "မင်္ဂလာပါ! ကျွန်ုပ်ကတော့ သင့်ရဲ့ ကိုယ်ပိုင် ငွေကြေးစီမံခန့်ခွဲမှု ဘော့ပါ။ အောက်ပါ ခလုတ်များဖြင့် စတင်အသုံးပြုနိုင်ပါပြီ။",
    "main_reply_buttons": [
        ["💰 ဝင်ငွေ", "💸 ထွက်ငွေ"],
        ["📊 အခြေအနေ", "🧾 အစီရင်ခံစာ"],
        ["🗓️ Reminder", "🎯 ဘတ်ဂျက်"],
        ["⚙️ စီမံခန့်ခွဲ", "⭐️ Premium"],
        ["💡 အကြံပြုမည်", "🔒 Privacy"]
    ],
    "help_text": "ကျွန်ုပ်သည် သင့်ရဲ့ ဝင်ငွေနှင့် အသုံးစရိတ်များကို စနစ်တကျမှတ်သားပြီး၊ အစီရင်ခံစာများ ထုတ်ပေးနိုင်သော ဘော့ဖြစ်ပါသည်။\n\n**အဓိက လုပ်ဆောင်ချက်များ:**\n- `[ဝင်ငွေ/ထွက်ငွေ] [ပမာဏ] [ဖော်ပြချက်]` ပုံစံဖြင့် ငွေစာရင်းသွင်းပါ။\n- `📊 အခြေအနေကြည့်ရန်`: လက်ရှိလ စုစုပေါင်းအခြေအနေကြည့်ရန်\n- `🧾 လစဉ်အစီရင်ခံစာ`: လစဉ်အစီရင်ခံစာ (PDF/Excel) ထုတ်ရန်\n- `/budget_status`: လစဉ်ဘတ်ဂျက်အခြေအနေကြည့်ရန်\n- `⭐️ Premium Features`: အခပေး လုပ်ဆောင်ချက်များ စုံစမ်းရန်\n- `/help`: အကူအညီ ရယူရန်\n- `/admin`: (Admin Only) Admin Dashboard ကို ဖွင့်ရန်\n- `🔒 Privacy`: လုံခြုံရေးနှင့် ပုဂ္ဂိုလ်ရေးဆိုင်ရာ မူဝါဒများ\n- `/delete_my_data`: သင်၏ အချက်အလက်များအားလုံးကို ဖျက်ရန်",
    "start_add_income": "💰 ဝင်ငွေ မှတ်တမ်းတင်ရန်၊ အောက်ပါပုံစံဖြင့် ရိုက်ထည့်ပေးပါ။\n\n`ဝင်ငွေ [ပမာဏ] [ဖော်ပြချက်]`\nဥပမာ။ ။ `ဝင်ငွေ 500000 လစာ`",
    "start_add_expense": "💸 ထွက်ငွေ မှတ်တမ်းတင်ရန်၊ အောက်ပါပုံစံဖြင့် ရိုက်ထည့်ပေးပါ။\n\n`ထွက်ငွေ [ပမာဏ] [ဖော်ပြချက်]`\nဥပမာ။ ။ `ထွက်ငွေ 15000 စားသောက်စရိတ်`",
    "expense_categories": ["စားသောက်စရိတ်", "အိမ်ငှားခ/လစဉ်ဘေလ်", "ပို့ဆောင်ရေး", "ဖျော်ဖြေရေး", "အဝတ်အထည်", "ကျန်းမာရေး", "အခြား"],
    "income_categories": ["လစာ", "ဘောနပ်စ်", "ရောင်းရငွေ", "ရင်းနှီးမြှုပ်နှံမှု", "အခြား"],
    "invalid_format": "❌ ပုံစံမမှန်ကန်ပါ။ ကျေးဇူးပြု၍ `ဝင်ငွေ/ထွက်ငွေ [ပမာဏ] [ဖော်ပြချက်]` ပုံစံဖြင့် ရိုက်ထည့်ပေးပါ။ ပမာဏမှာ ဂဏန်းသာ ဖြစ်ရပါမည်။",
    "data_saved": "✅ {category} အတွက် {amount:,.0f} Ks ကို မှတ်တမ်းတင်လိုက်ပါပြီ။",
    "summary_current_month": "{month} အတွက် စုစုပေါင်း ငွေကြေးအခြေအနေ:",
    "summary_details": "📈 စုစုပေါင်း ဝင်ငွေ: {income:,.0f} Ks\n📉 စုစုပေါင်း ထွက်ငွေ: {expense:,.0f} Ks\n💵 လက်ကျန်ငွေ: {balance:,.0f} Ks",
    "no_data": "ℹ️ မှတ်တမ်းတင်ထားသော ငွေကြေးအချက်အလက် မရှိသေးပါ။",
    "export_select_month": "အစီရင်ခံစာ ထုတ်ယူလိုသော လကို ရွေးချယ်ပါ။",
    "export_select_type": "အစီရင်ခံစာအတွက် File အမျိုးအစားကို ရွေးချယ်ပါ။",
    "export_success": "✅ {month} အတွက် {type} အစီရင်ခံစာကို အောက်ပါအတိုင်း ထုတ်ယူပေးလိုက်ပါပြီ။",
    "export_failure": "❌ အစီရင်ခံစာ ထုတ်ယူရာတွင် အမှားဖြစ်ပွားပါသည်။",
    "budget_set_start": "လစဉ် ဘတ်ဂျက် သတ်မှတ်ရန်၊ အောက်ပါပုံစံဖြင့် ရိုက်ထည့်ပေးပါ။\n\n`ဘတ်ဂျက် [ပမာဏ] [Category]`\nဥပမာ။ ။ `ဘတ်ဂျက် 300000 စားသောက်စရိတ်`",
    "budget_set_success": "✅ {category} အတွက် လစဉ် ဘတ်ဂျက် {amount:,.0f} Ks ကို သတ်မှတ်လိုက်ပါပြီ။",
    "budget_status_details": "💰 {month} အတွက် ဘတ်ဂျက်အခြေအနေ:\n\n{details}",
    "budget_no_set": "ℹ️ ဘတ်ဂျက် သတ်မှတ်ထားခြင်း မရှိသေးပါ။",
    "reminder_set_start": "🗓️ **သတိပေးချက် စီမံခန့်ခွဲခြင်း**\n\nသင်လိုချင်သော သတိပေးချက် အမျိုးအစားကို ဖွင့်/ပိတ် လုပ်နိုင်ပါသည်။",
    "reminder_set_success": "✅ အပတ်စဉ် အစီရင်ခံစာကို {day} နေ့တိုင်း ပို့ပေးပါမည်။",
    "unknown_command": "❌ နားမလည်သော စာသား သို့မဟုတ် command ဖြစ်ပါသည်။ /start ကို နှိပ်၍ အစမှ ပြန်စပါ။",
    "data_not_found": "မှတ်တမ်းတင်ထားသော အချက်အလက် မရှိပါ။",
    "processing": "အစီရင်ခံစာကို ပြုလုပ်နေပါသည်။",
    "premium_menu_header": "⭐️ **Premium Features များ**",
    "premium_menu_content": "Premium အသုံးပြုသူများအတွက်သာ ရရှိနိုင်သော လုပ်ဆောင်ချက်များ:\n\n1. 📊 **အသေးစိတ်သုံးစွဲမှုခွဲခြမ်းစိတ်ဖြာချက်** (Charts ပါဝင်)\n2. 🏷️ **စိတ်ကြိုက် Category များ ဖန်တီးခြင်း**\n3. 📄 **အဆင့်မြင့် အစီရင်ခံစာများ** (PDF Chart / Custom Date)\n4. 🎯 **ငွေကြေးပန်းတိုင်များ သတ်မှတ်ခြင်း**\n5. 🔔 **အသိဉာဏ်မြင့် သတိပေးချက်များ** (Budget Alert, Daily Reminder)\n6. 🔁 **လစဉ် ထပ်တလဲလဲ ငွေစာရင်း** (Recurring Transactions)\n\nသင့်ရဲ့ ဘဏ္ဍာရေးကို နောက်တစ်ဆင့် တက်လှမ်းလိုက်ပါ။",

    "premium_paywall": "🚫 **Premium Feature သုံးစွဲခွင့် မရှိသေးပါ** 🚫\n\nဤလုပ်ဆောင်ချက်သည် Premium အသုံးပြုသူများအတွက်သာ ဖြစ်ပါသည်။\n\nPremium Plan ဝယ်ယူရန် 'Premium Plan ယူရန်' ကို နှိပ်ပါ သို့မဟုတ် ၇ ရက် အစမ်းသုံးရန် 'Free Trial' ကို နှိပ်ပါ။",

    "premium_select_duration": "⭐️ **Premium Plan ရွေးချယ်ရန်**\n\nသင်ဝယ်ယူလိုသော လ/နှစ် ကို ရွေးချယ်ပေးပါ။",
    "premium_select_payment": "💳 **ငွေပေးချေမှု ရွေးချယ်ရန်**\n\n**{duration_text}** Plan အတွက် **{price_text}** ကျသင့်ပါမည်။\n\nငွေလွှဲလိုသော Payment Method ကို ရွေးချယ်ပေးပါ။",

    "premium_duration_1": "၁ လ",
    "premium_duration_6": "၆ လ (၁၀% လျှော့ဈေး)",
    "premium_duration_12": "၁ နှစ် (၂၀% လျှော့ဈေး)",

    "premium_price_1": "၅,၀၀၀ Ks",
    "premium_price_6": "၂၇,၀၀၀ Ks (မူရင်း ၃၀,၀၀၀ Ks)",
    "premium_price_12": "၄၈,၀၀၀ Ks (မူရင်း ၆၀,၀၀၀ Ks)",

    "premium_payment_wave": "WavePay",
    "premium_payment_kpay": "KBZPay",
    "premium_payment_aya": "AYA Bank",

    "premium_payment_details_wave": "အမည်: `U Aung Thu`\nဖုန်းနံပါတ်: `09 676 587 798`",
    "premium_payment_details_kpay": "အမည်: `U Aung Thu`\nဖုန်းနံပါတ်: `09 770 898 468`",
    "premium_payment_details_aya": "အမည်: `U Aung Thu`\nBank Account: `1234 5678 9012`",

    "premium_granted": "🎉 **ဂုဏ်ယူပါတယ်!** သင်၏ Premium သုံးစွဲခွင့်ကို {end_date} နေ့အထိ ဖွင့်ပေးလိုက်ပါပြီ။ Premium လုပ်ဆောင်ချက်အားလုံးကို စတင်အသုံးပြုနိုင်ပါပြီ။",
    "premium_trial_granted": "🎁 **Free Trial စတင်ပါပြီ!**\nPremium လုပ်ဆောင်ချက်အားလုံးကို ၇ ရက် ( {end_date} နေ့အထိ ) အခမဲ့ သုံးစွဲနိုင်ပါပြီ။ မပြည့်ခင် အခပေး Plan သို့ ပြောင်းလဲနိုင်ပါသည်။",

    # --- (STEP 3) Admin Approval System Texts ---
    "premium_final_instructions_admin_link_removed": """{payment_title} ဖြင့် ငွေလွှဲရန် အချက်အလက်
----------------------------------
{payment_details}
----------------------------------

**ကျေးဇူးပြု၍**
၁။ အထက်ပါ {payment_method} အကောင့်သို့ **{price_text}** လွှဲပေးပါ။
၂။ ငွေလွှဲပြီးပါက အောက်ပါ **"ငွေလွှဲပြီးပါပြီ"** ခလုတ်ကို နှိပ်ပါ။
၃။ ထို့နောက် ငွေလွှဲ Screenshot ကို ဤ chat ထဲသို့ ပို့ပေးပါ။

""",
    "premium_payment_complete_button": "✅ ငွေလွှဲပြီးပါပြီ (Screenshot ပို့ရန်)",
    "premium_awaiting_screenshot": "📸 ကျေးဇူးပြု၍ သင်၏ ငွေလွှဲ Screenshot ကို **ဤ chat ထဲသို့** အခု ပို့ပေးပါ။\n\nAdmin မှ စစ်ဆေးပြီး premium ကို ဖွင့်ပေးပါမည်။",
    "premium_screenshot_received": "✅ သင်၏ Screenshot ကို Admin ထံသို့ ပို့လိုက်ပါပြီ။\n\nကျေးဇူးပြု၍ ခဏစောင့်ပါ။ Admin မှ စစ်ဆေးပြီး အတည်ပြုပေးပါမည်။",
    "admin_approval_message": """⭐️ **Premium Approval Request**
----------------------------------
**User:** {user_name}
**User ID:** `{user_id}`
**Plan:** {duration_text} ({price_text})
----------------------------------
(အထက်ပါ Screenshot ကို စစ်ဆေးပြီး အတည်ပြုပေးပါ)
""",
    "admin_approval_granted": "✅ Approved!\nUser {user_id} ({user_name}) ကို {days} ရက် Premium ဖွင့်ပေးလိုက်ပါပြီ။",
    "admin_approval_rejected": "❌ Rejected!\nUser {user_id} ({user_name}) ကို Premium ပယ်ချကြောင်း အကြောင်းကြားလိုက်ပါပြီ။",
    "user_approval_rejected": "❌ သင်၏ Premium ဝယ်ယူမှုကို Admin မှ ပယ်ချလိုက်ပါသည်။\n\nငွေလွှဲ Screenshot မမှန်ကန်ပါက ထပ်မံ ကြိုးစားနိုင်ပါသည်။ အသေးစိတ် သိရှိလိုပါက Admin (@adu1010101) ထံ ဆက်သွယ်ပါ။",

    "custom_category_menu": "🏷️ **စိတ်ကြိုက် Category စီမံခန့်ခွဲခြင်း**\n\nသင်လိုချင်သည့် ဝင်ငွေ/ထွက်ငွေ အမျိုးအစား အသစ်များကို ဖန်တီးနိုင်ပါသည်။",
    "select_cat_type": "ဝင်ငွေ (သို့) ထွက်ငွေ အတွက် Category အမျိုးအစားကို ရွေးချယ်ပါ။",
    "add_cat_prompt": "ထည့်သွင်းလိုသော Category အမည်ကို ရိုက်ထည့်ပေးပါ။\n\nဥပမာ။ ။ `အိမ်မွေးတိရစ္ဆာန် စရိတ်`",
    "remove_cat_prompt": "ဖယ်ရှားလိုသော Category ကို ရွေးချယ်ပါ (မူရင်း Category များ မပါဝင်ပါ)။",
    "cat_add_success": "✅ '{name}' Category ကို ထည့်သွင်းပြီးပါပြီ။",
    "cat_add_fail_exists": "❌ '{name}' Category သည် ရှိနှင့်ပြီးသား သို့မဟုတ် မူရင်း Category တစ်ခု ဖြစ်နေပါသည်။",
    "cat_remove_success": "🗑️ '{name}' Category ကို ဖယ်ရှားပြီးပါပြီ။",
    "cat_remove_fail": "❌ '{name}' Category ကို ရှာမတွေ့ပါ သို့မဟုတ် ဖယ်ရှားမရပါ။",
    "custom_report_start": "📄 **ရက်သတ်မှတ်ချက်ဖြင့် အစီရင်ခံစာ ထုတ်ယူခြင်း**",
    "custom_report_prompt_start_date": "အစီရင်ခံစာ စတင်လိုသော **ရက်စွဲ (Start Date)** ကို `MM/DD/YYYY` ပုံစံဖြင့် ရိုက်ထည့်ပေးပါ။\n\nဥပမာ။ ။ `10/01/2024`",
    "custom_report_prompt_end_date": "အစီရင်ခံစာ ပြီးဆုံးလိုသော **ရက်စွဲ (End Date)** ကို `MM/DD/YYYY` ပုံစံဖြင့် ရိုက်ထည့်ပေးပါ။\n\nဥပမာ။ ။ `10/31/2024`",
    "custom_report_invalid_date": "❌ ရက်စွဲပုံစံ မမှန်ကန်ပါ။ `MM/DD/YYYY` ပုံစံ (ဥပမာ။ ။ 10/15/2024) ဖြင့်သာ ရိုက်ထည့်ပေးပါ။",
    "manage_tx_menu": "⚙️ **ငွေကြေးမှတ်တမ်း စီမံခန့်ခွဲခြင်း**\n\nသင်၏ ငွေစာရင်းများကို စီမံခန့်ခွဲနိုင်ပါသည်။",
    "no_recent_tx": "ℹ️ မကြာသေးမီက မှတ်တမ်းတင်ထားသော ငွေစာရင်း မရှိသေးပါ။",
    "select_tx_action": "👇 ဘယ်မှတ်တမ်းကို စီမံခန့်ခွဲမလဲ ရွေးချယ်ပေးပါ။",
    "tx_delete_success": "🗑️ မှတ်တမ်းကို ဖျက်လိုက်ပါပြီ။",
    "tx_edit_prompt": "✏️ **မှတ်တမ်း ပြင်ဆင်ခြင်း**\n\n `{date} - {type} ({category}) : {amount:,.0f} Ks` မှတ်တမ်းကို ဘယ်လို ပြင်ချင်ပါသလဲ?\n\nအောက်ပါ ပုံစံဖြင့် အသစ်ပြန်ရိုက်ပေးပါ။\n\n`[ဝင်ငွေ/ထွက်ငွေ] [ပမာဏ] [ဖော်ပြချက်]`\nဥပမာ။ ။ `ထွက်ငွေ 12000 နေ့လယ်စာ`",
    "tx_edit_success": "✅ မှတ်တမ်းအဟောင်းကို `{new_type} ({new_category}) : {new_amount:,.0f} Ks` ဖြင့် အောင်မြင်စွာ ပြင်ဆင်လိုက်ပါပြီ။",
    "tx_not_found": "❌ မှတ်တမ်းကို ရှာမတွေ့ပါ။",

    "goal_menu_header": "🎯 **ငွေကြေး ပန်းတိုင်များ စီမံခန့်ခွဲခြင်း**\n\nသင်၏ ငွေကြေး ပန်းတိုင်များကို ထားရှိပြီး စုဆောင်းမှု အခြေအနေ (Progress) ကို ခြေရာခံနိုင်ပါသည်။",
    "goal_add_prompt": "🎯 ပန်းတိုင်အသစ် ထည့်သွင်းရန်၊ အောက်ပါပုံစံဖြင့် ရိုက်ထည့်ပေးပါ။\n\n`[ပန်းတိုင်အမည်] [ပန်းတိုင်ပမာဏ] [ရက်စွဲ (MM/DD/YYYY)]`\nဥပမာ။ ။ `ဖုန်းအသစ် 500000 12/31/2025`",
    "goal_invalid_format": "❌ ပုံစံမမှန်ကန်ပါ။ `[ပန်းတိုင်အမည်] [ပန်းတိုင်ပမာဏ] [ရက်စွဲ (MM/DD/YYYY)]` ပုံစံဖြင့် ထည့်သွင်းပါ။ ပမာဏသည် ဂဏန်းဖြစ်ရပြီး ရက်စွဲသည် MM/DD/YYYY ပုံစံဖြစ်ရပါမည်။",
    "goal_add_success": "✅ '{name}' Goal ကို {amount:,.0f} Ks ဖြင့် {date} အတွက် သတ်မှတ်လိုက်ပါပြီ။",
    "goal_no_set": "ℹ️ သတ်မှတ်ထားသော ငွေကြေး ပန်းတိုင်များ မရှိသေးပါ။",
    "goal_progress_header": "🎯 **ငွေကြေး ပန်းတိုင်များ အခြေအနေ**",
    "goal_progress_detail": "\n\n{emoji} **{name}**\n*ရည်မှန်းချက်:* {amount:,.0f} Ks ( {date} )\n*စုဆောင်းပြီး:* {current_savings:,.0f} Ks\n*လိုအပ်ချက်:* {remaining:,.0f} Ks\n*Progress:* {progress:.1f}%",
    "goal_delete_menu": "🗑️ ဖျက်ပစ်လိုသော ပန်းတိုင်ကို ရွေးချယ်ပါ။",
    "goal_delete_success": "✅ '{name}' ပန်းတိုင်ကို ဖျက်လိုက်ပါပြီ။",
    "goal_not_found": "❌ ပန်းတိုင်ကို ရှာမတွေ့ပါ။",

    "budget_alert_overrun": "⚠️ **ဘတ်ဂျက် သတိပေးချက်!**\n\n'{category}' အတွက် သတ်မှတ်ဘတ်ဂျက် {budget:,.0f} Ks ၏ **{percent:.0f}%** ({spent:,.0f} Ks) ကို သုံးစွဲပြီးသွားပါပြီ။ ကျန်ရှိငွေ {remaining:,.0f} Ks သာ ရှိပါတော့သည်။",
    "budget_daily_spending": "\n\n**🎯 နေ့စဉ် ပျမ်းမျှ သုံးစွဲခွင့် (Daily Spending Limit)**\n*ကျန်ရှိရက်:* {days_remaining} ရက်\n*ပျမ်းမျှ သုံးစွဲခွင့်:* **{daily_limit:,.0f} Ks / တစ်ရက်**",

    "daily_reminder_morning": "☀️ **မနက်ခင်း ငွေစာရင်း သတိပေးချက်**\n\nယနေ့အတွက် ဝင်ငွေ/ထွက်ငွေ မှတ်တမ်းများ ထည့်သွင်းရန် မမေ့ပါနဲ့နော်။ နေ့လယ်စာ ကုန်ကျစရိတ်တွေကို မှတ်တမ်းတင်ဖို့ အဆင်သင့်ပြင်ထားလိုက်ပါ။",
    "daily_reminder_evening": "🌙 **ညနေခင်း ငွေစာရင်း သတိပေးချက်**\n\nယနေ့ တစ်နေကုန် သုံးစွဲခဲ့သော ငွေစာရင်း မှတ်တမ်းများကို အပြီးသတ် စစ်ဆေး မှတ်တမ်းတင်ပြီးပြီလား?\n\n💰 ဝင်ငွေ သို့ 💸 ထွက်ငွေ ကို ရိုက်ထည့်၍ စတင်နိုင်ပါသည်။",
    "daily_reminder_on": "✅ နေ့စဉ် သုံးစွဲမှု သတိပေးချက်ကို **နံနက် ၉ နာရီနှင့် ည ၇ နာရီ** တွင် စတင် ပို့ပေးပါမည်။",
    "daily_reminder_off": "❌ နေ့စဉ် သုံးစွဲမှု သတိပေးချက်ကို ပိတ်လိုက်ပါပြီ။",

    "privacy_policy": "🔒 **လုံခြုံရေးနှင့် ပုဂ္ဂိုလ်ရေးဆိုင်ရာ မူဝါဒ (/privacy)**\n\n၁။ **အချက်အလက် သိမ်းဆည်းခြင်း**:\nကျွန်ုပ်တို့သည် သင်၏ ငွေကြေးမှတ်တမ်းများ (transactions)၊ ဘတ်ဂျက်များ၊ နှင့် Premium အခြေအနေများကိုသာ Database တွင် သိမ်းဆည်းပါသည်။ သင်၏ Telegram အကောင့်အချက်အလက် (ဥပမာ- ဖုန်းနံပါတ်) ကို သိမ်းဆည်းထားခြင်း မရှိပါ။\n\n၂။ **အချက်အလက် လုံခြုံရေး (Data Security)**:\nသင်၏ အချက်အလက်အားလုံးကို ကျွန်ုပ်တို့၏ server ပေါ်ရှိ Database file (`financebot.db`) ထဲတွင် သိမ်းဆည်းထားပါသည်။\n\n၃။ **အချက်အလက် မျှဝေခြင်း**:\nသင်၏ ငွေကြေးဆိုင်ရာ အချက်အလက်များကို မည်သည့် တတိယပုဂ္ဂိုလ် (third-party) ကိုမျှ မျှဝေခြင်း၊ ရောင်းချခြင်း ပြုလုပ်မည် မဟုတ်ပါ။\n\n၄။ **အချက်အလက် ဖျက်ပစ်ခြင်း**:\nသင်သည် သင်၏ အချက်အလက်များအားလုံးကို အပြီးတိုင် ဖျက်ပစ်လိုပါက `/delete_my_data` command ကို အသုံးပြုနိုင်ပါသည်။ ဤသို့ ဖျက်လိုက်သည်နှင့် data အားလုံးကို ပြန်လည်ရယူနိုင်တော့မည် မဟုတ်ပါ။",

    "delete_data_confirm": "⚠️ **အချက်အလက် ဖျက်သိမ်းမှု အတည်ပြုခြင်း**\n\nသင်၏ ငွေကြေးမှတ်တမ်းများ၊ ဘတ်ဂျက်များ၊ ပန်းတိုင်များ၊ နှင့် Premium အချက်အလက်များ အပါအဝင် ဤ bot တွင် သိမ်းဆည်းထားသော **သင်၏ data အားလုံးကို** အပြီးတိုင် ဖျက်ပစ်ပါမည်။\n\nဤလုပ်ဆောင်ချက်ကို **ပြန်ပြင်နိုင်မည် မဟုတ်ပါ (Cannot Undo)**။\n\nသင် တကယ် သေချာပါသလား?",
    "delete_data_final_confirm": "🗑️ ဟုတ်ကဲ့၊ ကျွန်ုပ်၏ data အားလုံးကို ဖျက်ပစ်ပါ။",
    "delete_data_cancel": "❌ မလုပ်တော့ပါ။",
    "delete_data_success": "✅ သင်၏ အချက်အလက်များအားလုံးကို အောင်မြင်စွာ ဖျက်ပစ်လိုက်ပါပြီ။ Bot ကို ပြန်လည်အသုံးပြုလိုပါက /start ကို နှိပ်ပါ။",
    "delete_data_cancelled": "ℹ️ အချက်အလက် ဖျက်သိမ်းခြင်းကို ပယ်ဖျက်လိုက်ပါသည်။",

    "recurring_tx_menu_header": "🔁 **လစဉ် ထပ်တလဲလဲ ငွေစာရင်း (Recurring Transactions)**\n\nလစဉ် ပုံမှန် ဝင်/ထွက်မည့် ငွေစာရင်းများကို သတ်မှတ်ထားနိုင်ပါသည်။ Bot မှ သတ်မှတ်ရက်ရောက်တိုင်း အလိုအလျောက် စာရင်းသွင်းပေးပါမည်။",
    "recurring_tx_add_prompt": "🔁 **Recurring Transaction အသစ်ထည့်ရန်**\n\nအောက်ပါပုံစံဖြင့် ရိုက်ထည့်ပေးပါ။\n`[ဝင်ငွေ/ထွက်ငွေ] [ပမာဏ] [ဖော်ပြချက်] [နေ့ (1-28)]`\n\nဥပမာ။ ။ `ဝင်ငွေ 500000 လစာ 25` (လစဉ် ၂၅ ရက်နေ့)\nဥပမာ။ ။ `ထွက်ငွေ 100000 အိမ်လခ 1` (လစဉ် ၁ ရက်နေ့)\n\n(မှတ်ချက်: လအဆုံးရက်များ ရှုပ်ထွေးမှုမရှိစေရန် ၁ ရက်မှ ၂၈ ရက်နေ့အထိသာ သတ်မှတ်နိုင်ပါသည်။)",
    "recurring_tx_invalid_format": "❌ ပုံစံမမှန်ကန်ပါ။ `[ဝင်ငွေ/ထွက်ငွေ] [ပမာဏ] [ဖော်ပြချက်] [နေ့ (1-28)]` ပုံစံဖြင့် ထည့်သွင်းပါ။ နေ့ရက်သည် ၁ မှ ၂၈ အတွင်း ဂဏန်းဖြစ်ရပါမည်။",
    "recurring_tx_add_success": "✅ '{desc}' ({amount:,.0f} Ks) ကို လစဉ် {day} ရက်နေ့တိုင်း အလိုအလျောက် စာရင်းသွင်းပါမည်။",
    "recurring_tx_no_set": "ℹ️ သတ်မှတ်ထားသော လစဉ် ထပ်တလဲလဲ ငွေစာရင်း မရှိသေးပါ။",
    "recurring_tx_delete_menu": "🗑️ ဖျက်ပစ်လိုသော လစဉ်ငွေစာရင်းကို ရွေးချယ်ပါ။",
    "recurring_tx_delete_success": "✅ '{name}' recurring transaction ကို ဖျက်လိုက်ပါပြီ။",
    "recurring_tx_not_found": "❌ Recurring transaction ကို ရှာမတွေ့ပါ။",
    "recurring_tx_executed": "🔁 **Recurring Transaction**\n\nသင်၏ လစဉ်စာရင်း '{desc}' ({amount:,.0f} Ks) ကို ယနေ့အတွက် အလိုအလျောက် မှတ်တမ်းတင်လိုက်ပါပြီ။",

    # --- (STEP 4) NEW: Admin Dashboard Texts ---
    "not_admin": "🚫 ဤ command ကို Admin သာ အသုံးပြုနိုင်ပါသည်။",
    "admin_dashboard_header": "⚙️ **Admin Control Panel** ⚙️\n\nကြိုဆိုပါတယ် Admin! အောက်ပါခလုတ်များဖြင့် bot ကို ထိန်းချုပ်နိုင်ပါပြီ။",
    "admin_stats_button": "📊 Statistics ကြည့်ရန်",
    "admin_broadcast_button": "📣 User အားလုံးထံ ကြေငြာချက်ပို့ရန်",
    "admin_find_user_button": "👤 User တစ်ယောက်ချင်း ရှာဖွေရန်",
    "admin_stats_message": "📊 **Bot Statistics**\n\n👥 **စုစုပေါင်း User:** {total} ယောက်\n⭐️ **Premium User:** {premium} ယောက်",
    "admin_broadcast_prompt": "📣 **Broadcast Mode**\n\nUser အားလုံးထံ ပို့လိုသော message ကို ရိုက်ထည့်ပေးပါ။ (Markdown/HTML သုံးနိုင်ပါသည်)။\n\nပယ်ဖျက်လိုပါက `cancel` ဟု ရိုက်ထည့်ပါ။",
    "admin_broadcast_confirm": "⚠️ **Broadcast Confirmation**\n\nအောက်ပါ message ကို User **{count}** ယောက်ထံ ပို့ပါမည်။\n----------------------------------\n{message}\n----------------------------------\n\nပို့ရန် သေချာပါသလား?",
    "admin_broadcast_confirm_button": "✅ ဟုတ်ကဲ့၊ ပို့ပါ။",
    "admin_broadcast_cancel_button": "❌ မပို့တော့ပါ။",
    "admin_broadcast_start": "⏳ Broadcast စတင်နေပါပြီ... User {count} ယောက်ထံ ပို့ပါမည်။ ပြီးဆုံးပါက အကြောင်းပြန်ပါမည်။",
    "admin_broadcast_complete": "✅ **Broadcast Complete!**\n\n- **အောင်မြင်:** {sent} ယောက်\n- **မအောင်မြင် (Bot ကို block သွားသူများ):** {failed} ယောက်",
    "admin_broadcast_cancelled": "❌ Broadcast ကို ပယ်ဖျက်လိုက်ပါသည်။",
    "admin_find_user_prompt": "👤 **Find User**\n\nရှာဖွေလိုသော User ၏ **Telegram User ID** (ဂဏန်း) ကို ရိုက်ထည့်ပေးပါ။\n\nပယ်ဖျက်လိုပါက `cancel` ဟု ရိုက်ထည့်ပါ။",
    "admin_user_not_found": "❌ User ID `{user_id}` ကို database ထဲတွင် ရှာမတွေ့ပါ။",
    "admin_user_details": """👤 **User Details**
----------------------------------
**User ID:** `{id}`
**Premium Status:** {status}
**Premium End Date:** {end_date}
**Total Transactions:** {tx_count}
**Used Trial:** {used_trial}
----------------------------------
👇 အောက်ပါခလုတ်များဖြင့် ထိန်းချုပ်နိုင်ပါသည်။
""",
    "admin_grant_button": "➕ Premium 30 ရက် ထပ်တိုးရန်",
    "admin_revoke_button": "➖ Premium ဖျက်သိမ်းရန်",
    "admin_user_granted": "✅ User {user_id} ကို Premium 30 ရက် ထပ်တိုးပေးလိုက်ပါပြီ။",
    "admin_user_revoked": "✅ User {user_id} ၏ Premium ကို ဖျက်သိမ်းလိုက်ပါပြီ။",

    # --- (STEP 5) NEW: Quick Add Texts ---
    "quick_add_prompt_type": "💰 **{amount:,.0f} Ks** ကို ဘာအဖြစ် မှတ်သားမလဲ။",
    "quick_add_prompt_category": "💸 **{amount:,.0f} Ks** ကို ထွက်ငွေ (Expense) အဖြစ် မှတ်သားပါမည်။\n\n👇 ကျေးဇူးပြု၍ Category တစ်ခု ရွေးချယ်ပါ။",
    "quick_add_type_income": "💰 ဝင်ငွေ",
    "quick_add_type_expense": "💸 ထွက်ငွေ",

    # --- (STEP 5.1) NEW: AI Financial Analyst Texts ---
    "ai_analysis_button": "💡 AI သုံးသပ်ချက်",
    "ai_analysis_header": "💡 **AI ဘဏ္ဍာရေး သုံးသပ်ချက်**\n\nသင့်ရဲ့ ပြီးခဲ့တဲ့ ရက် ၃၀ သုံးစွဲမှု ပုံစံကို AI က သုံးသပ်ပြီးပါပြီ-\n",
    "ai_analysis_no_data": "ℹ️ သုံးသပ်ချက် ပေးနိုင်ရန်အတွက် ပြီးခဲ့သော ရက် ၃၀ အတွင်း ဝင်ငွေ/ထွက်ငွေ မှတ်တမ်းများ လုံလောက်စွာ မရှိသေးပါ။",

    "ai_insight_saving_rate_low": "\n* **စုငွေ (Saving Rate):** သင်၏ စုငွေနှုန်းမှာ **{rate:.1f}%** သာ ဖြစ်နေပြီး၊ အကြံပြုထားသော (၁၅% - ၂၀%) ထက် နည်းနေပါသည်။ ထွက်ငွေများကို ပြန်လည်စစ်ဆေးသင့်ပါသည်။",
    "ai_insight_saving_rate_good": "\n* **စုငွေ (Saving Rate):** သင်၏ စုငွေနှုန်းမှာ **{rate:.1f}%** ဖြစ်ပြီး၊ အလွန်ကောင်းမွန်ပါသည်။ ဆက်လက် ထိန်းသိမ်းပါ။ 📈",
    "ai_insight_saving_rate_negative": "\n* **အရေးပေါ် (Negative Saving):** သင်၏ ထွက်ငွေ ( {expense:,.0f} Ks) သည် ဝင်ငွေ ( {income:,.0f} Ks) ထက် ပိုများနေပါသည်။ ငွေစာရင်းများကို ချက်ချင်း ပြန်လည် စစ်ဆေးပါ။ 🚨",

    "ai_insight_top_expense": "\n* **အဓိက သုံးစွဲမှု:** သင်၏ ထွက်ငွေ စုစုပေါင်း၏ **{percent:.0f}%** ကို **'{category}'** တွင် အများဆုံး သုံးစွဲထားပါသည်။",

    "ai_insight_budget_over": "\n* **ဘတ်ဂျက် ကျော်လွန်:** သင်၏ **'{category}'** ဘတ်ဂျက် ( {budget:,.0f} Ks) ကို **{percent:.0f}%** ( {spent:,.0f} Ks) အထိ သုံးစွဲထားပြီး၊ သတ်မှတ်ငွေထက် ကျော်လွန်နေပါပြီ။ ⚠️",
    "ai_insight_budget_warning": "\n* **ဘတ်ဂျက် သတိပေးချက်:** သင်၏ **'{category}'** ဘတ်ဂျက် ( {budget:,.0f} Ks) ကို **{percent:.0f}%** ( {spent:,.0f} Ks) အထိ သုံးစွဲထားပြီး၊ သတ်မှတ်ငွေ၏ ၈၀% နီးပါး ရောက်ရှိနေပါပြီ။",


    # --- (STEP 6) NEW: Backup/Restore Texts ---
    "backup_restore_button": "💾 Backup / Restore",
    "backup_restore_menu_header": "💾 **Data Backup & Restore**\n\nသင်၏ ငွေစာရင်း Data များကို JSON file အဖြစ် ထုတ်ယူ သိမ်းဆည်းနိုင် (Backup) ခြင်း၊ File မှ Data များကို ပြန်လည် ထည့်သွင်း (Restore) ခြင်း ပြုလုပ်နိုင်ပါသည်။",
    "backup_button": "📤 Backup ပြုလုပ်ရန် (Data ထုတ်ယူမည်)",
    "restore_button": "📥 Restore ပြန်လုပ်ရန် (Data ပြန်ထည့်မည်)",

    "backup_prompt_sending": "⏳ သင်၏ Data များကို စုစည်း၍ Backup file ပြုလုပ်နေပါသည်။ ခဏစောင့်ပါ။",
    "backup_prompt_success": "✅ သင်၏ Backup file (`backup_{date}.json`) ကို အောင်မြင်စွာ ထုတ်ယူပြီးပါပြီ။\n\nဤ file ကို လုံခြုံသော နေရာ (ဥပမာ- Email, Google Drive) တွင် သိမ်းဆည်းထားပါ။",

    "restore_prompt": "📥 **Data Restore ပြန်လုပ်ရန်**\n\nကျေးဇူးပြု၍ သင် သိမ်းဆည်းထားသော `backup_....json` file ကို ဤ chat ထဲသို့ ပို့ပေးပါ။\n\n**(!!!) သတိပြုရန်:** Restore ပြုလုပ်သည်နှင့် သင်၏ လက်ရှိ Data အားလုံး **ဖျက်သိမ်းခံရမည်** ဖြစ်ပြီး၊ ဤ file ထဲမှ Data များဖြင့် **အစားထိုး** သွားမည် ဖြစ်သည်။\n\nပယ်ဖျက်လိုပါက `cancel` ဟု ရိုက်ထည့်ပါ။",
    "restore_success": "✅ Data များကို Backup file မှ အောင်မြင်စွာ Restore ပြန်လုပ်ပြီးပါပြီ။",
    "restore_error_json": "❌ Error: ပို့လိုက်သော file သည် JSON file ပုံစံ မမှန်ကန်ပါ။",
    "restore_error_format": "❌ Error: ဤ file သည် ကျွန်ုပ်တို့၏ Backup file ပုံစံ မဟုတ်ပါ။ (Data များ မစုံလင်ပါ)",
    "restore_error_general": "❌ Restore ပြုလုပ်ရာတွင် အမှားအယွင်း ဖြစ်ပွားပါသည်။",
    "restore_cancelled": "❌ Restore ပြုလုပ်ခြင်းကို ပယ်ဖျက်လိုက်ပါသည်။",

    # ... (တခြား TEXTS တွေ)

    # (!!!) အောက်ပါ Code Block အသစ်ကို ထပ်ထည့်ပါ (!!!)
    # --- Generic Info Button Text ---
    "info_button_text": "ℹ️ အသုံးပြုပုံ", # <-- ခလုတ် နာမည်ကို တစ်နေရာတည်းမှာ ထိန်းချုပ်ရန်
    
    # --- Info Text (ရှင်းပြချက်များ) ---
    "info_analytics_text": "သင့်ရဲ့ ဝင်ငွေ/ထွက်ငွေ စာရင်းတွေကို Pie Chart, Bar Chart တွေနဲ့ အမြင်အာရုံဖြင့် ရှင်းလင်းစွာ ကြည့်ရှုနိုင်တဲ့ feature ပါ။",
    "info_ai_analysis_text": "သင့်ရဲ့ ဝင်ငွေ/ထွက်ငွေ ပုံစံကို AI က နေ့စဉ် သုံးသပ်ပြီး၊ ဘယ်နေရာမှာ ပိုသုံးနေလဲ၊ ဘယ်လို ချွေတာသင့်လဲ အကြံဉာဏ် ပေးမယ့် feature ပါ။",
    "info_custom_category_text": "ဒါက 'Gym ကစားခ', 'Pet Food' လိုမျိုး သင့်ရဲ့ ကိုယ်ပိုင် ဝင်ငွေ/ထွက်ငွေ Category အသစ်တွေ ထည့်/ဖျက် လုပ်နိုင်တဲ့ feature ပါ။",
    "info_custom_report_text": "သင်လိုချင်တဲ့ ရက်စွဲ (ဥပမာ- 10/01/2024 မှ 10/15/2024 အထိ) အတွင်းက ငွေစာရင်းတွေကိုပဲ သီးသန့် PDF/Excel ထုတ်ယူနိုင်တဲ့ feature ပါ။",
    "info_goal_tracking_text": "'ဖုန်းအသစ်ဝယ်ရန်' လိုမျိုး ပန်းတိုင်တွေ သတ်မှတ်ပြီး၊ ပိုက်ဆံ ဘယ်လောက် စုပြီးပြီလဲ၊ ဘယ်လောက် လိုသေးလဲ ခြေရာခံနိုင်တဲ့ feature ပါ။",
    "info_recurring_tx_text": "'အိမ်လခ', 'လစဉ် ဘေလ်' လိုမျိုး လတိုင်း ပုံမှန် ဝင်/ထွက် တဲ့ ငွေစာရင်းတွေကို bot က အလိုအလျောက် မှတ်သားပေးမယ့် feature ပါ။",
    "info_backup_restore_text": "သင့်ငွေစာရင်း Data တွေအားလုံးကို .json file အဖြစ် ထုတ်ယူ သိမ်းဆည်းထားနိုင်ပြီး၊ ဖုန်းပျောက်/အကောင့်ပြောင်း တဲ့အခါ Data တွေ ပြန်ထည့်နိုင်တဲ့ feature ပါ။",
    
    # --- Reminder Feature အတွက် အသစ် (၂) ခု ---
    "info_daily_reminder_text": "သင့်ရဲ့ နေ့စဉ် ငွေစာရင်း ထည့်သွင်းဖို့ မနက် ၉ နာရီ နှင့် ည ၇ နာရီ တွင် သတိပေးမယ့် feature ပါ။",
    "info_weekly_summary_text": "သင် သတ်မှတ်ထားတဲ့ နေ့ (ဥပမာ- တနင်္ဂနွေ) ရောက်တိုင်း၊ အဲ့ဒီ အပတ်ရဲ့ ငွေစာရင်း အကျဉ်းချုပ်ကို အလိုအလျောက် ပို့ပေးမယ့် feature ပါ။",
    "info_monthly_report_text": "သင်၏ လစဉ် ငွေစာရင်း အပြည့်အစုံကို PDF သို့မဟုတ် Excel file အဖြစ် ထုတ်ယူနိုင်ပါသည်။",
    "info_quick_add_text": "သင် ရိုက်ထည့်လိုက်သော ပမာဏကို 'ဝင်ငွေ' လား 'ထွက်ငွေ' လား ရွေးချယ်ပြီး အမြန် မှတ်တမ်းတင်နိုင်ပါသည်။",
    # --- Feedback Feature Texts ---
    "feedback_button": "💡 အကြံပြုမည်",
    "feedback_prompt": "ကျေးဇူးပြု၍ သင်၏ အကြံပြုချက်ကို ဤနေရာတွင် ရိုက်ထည့်ပေးပါ။\n\nပယ်ဖျက်လိုပါက `cancel` ဟု ရိုက်ထည့်ပါ။",
    "feedback_cancel": "❌ အကြံပြုချက် ပေးပို့ခြင်းကို ပယ်ဖျက်လိုက်ပါသည်။",
    "feedback_admin_header": "💡 **New Feedback Received!**",
    "feedback_admin_details": "----------------------------------\n**User:** {user_name}\n**User ID:** `{user_id}`\n**Feedback:**\n{feedback_text}",
    "feedback_success": "✅ သင်၏ အကြံပြုချက်ကို Admin ထံသို့ အောင်မြင်စွာ ပေးပို့လိုက်ပါပြီ။ ကျေးဇူးတင်ပါသည်။",
    "feedback_error": "❌ အမှားအယွင်းတစ်ခုကြောင့် Admin ထံ ပို့၍ မရသေးပါ။",
    # --- (!!!) အသစ်ထည့်ခြင်း ပြီးဆုံးပါပြီ (!!!) ---
    # ... (feedback_error Text ရဲ့ အောက်)
    
    # --- (!!!) NEW: Multi-Wallet Account Texts (!!!) ---
    "account_menu_header": "💰 **Account စီမံခန့်ခွဲခြင်း**\n\nသင်၏ ငွေစာရင်း Account များကို (ဥပမာ- Cash, KPay, Bank) ဤနေရာတွင် စီမံခန့်ခွဲနိုင်ပါသည်။",
    "account_add_button": "➕ Account အသစ်ထည့်ရန်",
    "account_view_button": "👀 Account များ ကြည့်ရန်",
    "account_transfer_button": "🔁 ငွေလွှဲပြောင်းရန် (Transfer)",
    "account_add_prompt": "🆕 **Account အသစ်ထည့်ရန်**\n\nAccount အမည်ကို ရိုက်ထည့်ပေးပါ။ (ဥပမာ- `Cash` သို့မဟုတ် `KPay`)\n\n(လက်ကျန်ငွေ ပါ ထည့်လိုပါက `[အမည်] [လက်ကျန်]` ဥပမာ- `Bank 100000`)\n\nပယ်ဖျက်လိုပါက `cancel` ဟု ရိုက်ထည့်ပါ။",
    "account_add_success": "✅ '{name}' Account ကို {balance:,.0f} Ks လက်ကျန်ဖြင့် အောင်မြင်စွာ ဖန်တီးပြီးပါပြီ။",
    "account_add_fail_exists": "❌ '{name}' အမည်ဖြင့် Account ရှိပြီးသားပါ။",
    # --- (!!!) NEW: Step 2 Account Texts (!!!) ---
    "select_account_prompt": "👇 **{tx_type} {amount:,.0f} Ks** ({desc}) ကို ဘယ် Account ထဲမှာ မှတ်တမ်းတင်မလဲ ရွေးချယ်ပေးပါ။",
    "select_account_button_none": "🔘 Account မသတ်မှတ် (Unassigned)",
    "data_saved_with_account": "✅ {category} အတွက် {amount:,.0f} Ks ကို **{account_name}** Account ထဲတွင် မှတ်တမ်းတင်လိုက်ပါပြီ။",
    "data_saved_no_account": "✅ {category} အတွက် {amount:,.0f} Ks ကို (Account မသတ်မှတ်) ဖြင့် မှတ်တမ်းတင်လိုက်ပါပြီ။",
    "no_accounts_error": "❌ သင့်တွင် Account မရှိသေးပါ။\n\nကျေးဇူးပြု၍ '⚙️ စီမံခန့်ခွဲ' -> '💰 Account စီမံခန့်ခွဲ' တွင် Account အနည်းဆုံး တစ်ခု အရင် ဖန်တီးပါ။ Account မဖန်တီးဘဲ သုံးစွဲ၍ မရတော့ပါ။",
    "tx_edit_prompt_account": "✏️ **Account ပြောင်းလဲရန်**\n\n`{desc}` ({amount:,.0f} Ks) အတွက် လက်ရှိ Account ({current_account}) မှ အောက်ပါ Account သို့ ပြောင်းလဲမည်-",
    "tx_edit_account_success": "✅ မှတ်တမ်း၏ Account ကို **{account_name}** သို့ ပြောင်းလဲလိုက်ပါပြီ။",
    # --- (!!!) End of New Texts (!!!) ---
    "account_list_header": "💰 **သင်၏ Account များ**\n",
    "account_list_detail": "\n- **{name}**: {balance:,.0f} Ks",
    "account_list_total": "\n\n**စုစုပေါင်း လက်ကျန် (Assigned):** {total:,.0f} Ks",
    "account_list_unassigned": "\n**Account မခွဲရသေးသော လက်ကျန်:** {unassigned:,.0f} Ks",
    "account_list_grand_total": "\n\n**Grand Total:** {total:,.0f} Ks",
    "account_list_empty": "ℹ️ သင့်တွင် Account များ ဖန်တီးထားခြင်း မရှိသေးပါ။ 'Account အသစ်ထည့်ရန်' ခလုတ်ကို နှိပ်ပြီး သင်၏ ပထမဆုံး Account ကို ဖန်တီးပါ။",
    # --- (!!!) End of New Texts (!!!) ---

    "premium_paywall": "🚫 **Premium Feature သုံးစွဲခွင့် မရှိသေးပါ** 🚫\n\nဤလုပ်ဆောင်ချက်သည် Premium အသုံးပြုသူများအတွက်သာ ဖြစ်ပါသည်။\n\nPremium Plan ဝယ်ယူရန် 'Premium Plan ယူရန်' ကို နှိပ်ပါ သို့မဟုတ် ၇ ရက် အစမ်းသုံးရန် 'Free Trial' ကို နှိပ်ပါ။",

}
# -------------------------------------------------------------------


# ----------------------- Helpers -----------------------
MYANMAR_MONTHS = {1: "ဇန်နဝါရီ", 2: "ဖေဖော်ဝါရီ", 3: "မတ်", 4: "ဧပြီ", 5: "မေ", 6: "ဇွန်",
                  7: "ဇူလိုင်", 8: "ဩဂုတ်", 9: "စက်တင်ဘာ", 10: "အောက်တိုဘာ", 11: "နိုဝင်ဘာ", 12: "ဒီဇင်ဘာ"}


def format_myanmar_date(date_obj):
    year = date_obj.year
    month = MYANMAR_MONTHS.get(date_obj.month, date_obj.strftime("%B"))
    return f"{year} ခုနှစ် {month} လ"

# ====================================================================
# ExportManager Class (NO CHANGES)
# ====================================================================


class ExportManager:
    def __init__(self, export_dir: str = EXPORT_DIR):
        self.export_dir = export_dir
        os.makedirs(self.export_dir, exist_ok=True)
        self.WEASYPRINT_AVAILABLE = WEASYPRINT_AVAILABLE

    def _validate_transactions(self, transactions: List[Dict[str, Any]]) -> pd.DataFrame:
        df = pd.DataFrame(transactions)
        if df.empty:
            return pd.DataFrame(columns=['Date', 'Type', 'description', 'Category', 'Income', 'Expense'])

        # Handle potential NaT/None in date column before conversion
        df = df.dropna(subset=['date'])
        df['Date'] = pd.to_datetime(df['date']).dt.strftime('%Y-%m-%d')

        df['Type'] = df['type'].apply(
            lambda x: 'ဝင်ငွေ' if x == 'income' else 'ထွက်ငွေ')
        df['Income'] = df.apply(lambda row: row['amount']
                                if row['type'] == 'income' else 0, axis=1)
        df['Expense'] = df.apply(
            lambda row: row['amount'] if row['type'] == 'expense' else 0, axis=1)
        df['Category'] = df['category']
        df['description'] = df['description'].fillna('')
        df['Income'] = df['Income'].astype(int)
        df['Expense'] = df['Expense'].astype(int)

        return df[['Date', 'Type', 'description', 'Category', 'Income', 'Expense']]

    def _export_weasyprint(self, title: str, transactions: List[Dict[str, Any]], chart_data: Optional[str] = None) -> Optional[io.BytesIO]:
        if not self.WEASYPRINT_AVAILABLE:
            return None
        df = self._validate_transactions(transactions)

        total_income = df['Income'].sum()
        total_expense = df['Expense'].sum()
        balance = total_income - total_expense

        font_style = 'font-family: "Pyidaungsu", "Noto Sans Myanmar", Padauk, sans-serif;'

        table_rows = ""
        for _, row in df.iterrows():
            description_safe = str(row['description']).replace(
                '&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

            table_rows += f"""
            <tr>
                <td>{row['Date']}</td>
                <td>{row['Type']}</td>
                <td>{description_safe}</td>
                <td>{row['Category']}</td>
                <td class="amount">{f"{row['Income']:,.0f}" if row['Income'] > 0 else "-"}</td>
                <td class="amount">{f"{row['Expense']:,.0f}" if row['Expense'] > 0 else "-"}</td>
            </tr>
            """

        chart_html = ""
        if chart_data:
            chart_html = f"""
            <div class="chart-container">
                <h2>အသုံးစရိတ် ခွဲခြမ်းစိတ်ဖြာချက် (Expense Analysis)</h2>
                <img src="{chart_data}" alt="Expense Pie Chart">
            </div>
            """

        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                @page {{ size: A4; margin: 1.5cm; }}
                @font-face {{
                    font-family: 'Pyidaungsu';
                    src: url("file:///{CLEAN_FONT_PATH}"); 
                }}
                
                body {{
                    {font_style} 
                    font-size: 10pt;
                    color: #333;
                }}
                .header {{
                    text-align: center;
                    margin-bottom: 1.5cm;
                    border-bottom: 2px solid #004a99;
                    padding-bottom: 10px;
                }}
                h1 {{
                    font-size: 18pt;
                    color: #004a99;
                    margin: 0;
                }}
                h2 {{
                    font-size: 14pt;
                    color: #333;
                    border-bottom: 1px solid #ccc;
                    padding-bottom: 5px;
                }}
                .chart-container {{
                    text-align: center;
                    margin-bottom: 1cm;
                }}
                .chart-container img {{
                    max-width: 80%;
                    height: auto;
                    margin-top: 10px;
                }}
                table {{
                    width: 100%; border-collapse: collapse; margin-top: 10px;
                }}
                th, td {{
                    border: 1px solid #ddd; padding: 8px; {font_style}
                    text-align: left;
                }}
                th {{
                    background-color: #f2f2f2; 
                    font-weight: bold;
                    color: #004a99;
                }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                .amount {{ text-align: right; }}
                
                .summary-section {{
                    margin-top: 1.5cm;
                    page-break-inside: avoid;
                }}
                .summary-container {{ 
                    width: 100%;
                    display: table;
                }}
                .summary-box {{
                    display: table-cell;
                    width: 30%;
                    padding: 15px;
                    border-radius: 5px;
                    text-align: right;
                    font-size: 11pt;
                }}
                .summary-income {{ background-color: #e6f7ff; border: 1px solid #b3e0ff; }}
                .summary-expense {{ background-color: #fff0f0; border: 1px solid #ffcccc; }}
                .summary-balance {{ background-color: #f0fff0; border: 1px solid #ccffcc; }}
                
                .summary-box .label {{ font-weight: bold; display: block; margin-bottom: 5px; }}
                .summary-box .value {{ font-size: 13pt; font-weight: bold; }}
                .summary-income .value {{ color: #0056b3; }}
                .summary-expense .value {{ color: #d90000; }}
                .summary-balance .value {{ color: #006400; }}
                
                .spacer {{ display: table-cell; width: 5%; }}
                
            </style>
        </head>
        <body>
            <div class="header">
                <h1>{title}</h1>
            </div>
            
            {chart_html}

            <h2>ငွေစာရင်း အသေးစိတ် (Transaction Details)</h2>
            <table>
                <thead>
                    <tr>
                        <th width="12%">ရက်စွဲ</th>
                        <th width="10%">အမျိုးအစား</th>
                        <th width="30%">ဖော်ပြချက်</th> 
                        <th width="18%">Category</th> 
                        <th width="15%" class="amount">ဝင်ငွေ (Ks)</th>
                        <th width="15%" class="amount">ထွက်ငွေ (Ks)</th>
                    </tr>
                </thead>
                <tbody>{table_rows}</tbody>
            </table>
            
            <div class="summary-section">
                <h2>အနှစ်ချုပ် (Summary)</h2>
                <div class="summary-container">
                    <div class="summary-box summary-income">
                        <span class="label">စုစုပေါင်း ဝင်ငွေ</span>
                        <span class="value">{total_income:,.0f} Ks</span>
                    </div>
                    <div class="spacer"></div>
                    <div class="summary-box summary-expense">
                        <span class="label">စုစုပေါင်း ထွက်ငွေ</span>
                        <span class="value">{total_expense:,.0f} Ks</span>
                    </div>
                    <div class="spacer"></div>
                    <div class="summary-box summary-balance">
                        <span class="label">လက်ကျန်ငွေ</span>
                        <span class="value">{balance:,.0f} Ks</span>
                    </div>
                </div>
            </div>
        </body>
        </html>
        """
        try:
            buffer = io.BytesIO()
            HTML(string=html_content).write_pdf(buffer)
            buffer.seek(0)
            return buffer
        except Exception as e:
            logger.error(f"Error generating PDF with WeasyPrint: {e}")
            return None

    def _export_to_excel(self, transactions: List[Dict[str, Any]]) -> Optional[io.BytesIO]:
        if not OPENPYXL_AVAILABLE:
            return None
        df = self._validate_transactions(transactions)

        buffer = io.BytesIO()
        try:
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Summary', index=False)
                sheet = writer.sheets['Summary']
                myanmar_font = Font(name='Pyidaungsu', size=10, bold=True)
                header_fill = PatternFill(
                    start_color="D3D3D3", fill_type="solid")
                center_alignment = Alignment(
                    horizontal='center', vertical='center')
                right_alignment = Alignment(
                    horizontal='right', vertical='center')

                for col_idx, col_name in enumerate(df.columns):
                    cell = sheet.cell(row=1, column=col_idx + 1)
                    cell.font = myanmar_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    sheet.column_dimensions[chr(
                        65 + col_idx)].width = max(len(col_name) * 1.5, 18)

                for row_idx in range(2, len(df) + 2):
                    sheet.cell(
                        row=row_idx, column=5).alignment = right_alignment
                    sheet.cell(
                        row=row_idx, column=6).alignment = right_alignment

                next_row = len(df) + 3
                income_total = df['Income'].sum()
                expense_total = df['Expense'].sum()
                balance = income_total - expense_total

                sheet.cell(row=next_row, column=4,
                           value="စုစုပေါင်း ဝင်ငွေ (Total Income):").font = myanmar_font
                sheet.cell(row=next_row, column=5,
                           value=income_total).font = myanmar_font
                sheet.cell(row=next_row, column=5).alignment = right_alignment

                sheet.cell(row=next_row + 1, column=4,
                           value="စုစုပေါင်း ထွက်ငွေ (Total Expense):").font = myanmar_font
                sheet.cell(row=next_row + 1, column=5,
                           value=expense_total).font = myanmar_font
                sheet.cell(row=next_row + 1,
                           column=5).alignment = right_alignment

                sheet.cell(row=next_row + 2, column=4,
                           value="လက်ကျန်ငွေ (Balance):").font = myanmar_font
                sheet.cell(row=next_row + 2, column=5,
                           value=balance).font = myanmar_font
                sheet.cell(row=next_row + 2,
                           column=5).alignment = right_alignment

            buffer.seek(0)
            return buffer
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            return None

    def export_data(self, title: str, transactions: List[Dict[str, Any]], export_type: str, chart_data: Optional[str] = None) -> Optional[io.BytesIO]:
        if not transactions:
            return None
        if export_type == 'pdf' and self.WEASYPRINT_AVAILABLE:
            return self._export_weasyprint(title, transactions, chart_data)
        elif export_type == 'xlsx' and OPENPYXL_AVAILABLE:
            return self._export_to_excel(transactions)
        return None


# ====================================================================
# PlotlyChartManager Class (NO CHANGES)
# ====================================================================
class PlotlyChartManager:
    def __init__(self):
        self.PLOTLY_AVAILABLE = PLOTLY_AVAILABLE
        self.WEASYPRINT_AVAILABLE = WEASYPRINT_AVAILABLE
        self.KALEIDO_AVAILABLE = KALEIDO_AVAILABLE
        self.myanmar_font_family = "Pyidaungsu, Noto Sans Myanmar, Padauk, sans-serif"

    def _render_plotly_fig_to_png(self, fig, width=800, height=600) -> Optional[io.BytesIO]:
        if not self.PLOTLY_AVAILABLE:
            return None
        # WeasyPrint ကို မသုံးတော့ဘဲ Kaleido တစ်ခုတည်းကိုပဲ သုံးပါမည်။
        if self.KALEIDO_AVAILABLE:
            try:
                buffer = io.BytesIO(pio.to_image(
                    fig, format='png', width=width, height=height))
                buffer.seek(0)
                logger.info("✅ Plotly Chart rendered to PNG using Kaleido.")
                return buffer
            except Exception as e:
                logger.error(f"Kaleido failed to render Plotly to PNG: {e}")
        
        # အပေါ်က Kaleido ပါ မအောင်မြင်ရင် Error ပြပါ
        logger.error(
            "❌ Failed to render Plotly chart to PNG. Kaleido is not working.")
        return None

        if self.KALEIDO_AVAILABLE:
            try:
                buffer = io.BytesIO(pio.to_image(
                    fig, format='png', width=width, height=height))
                buffer.seek(0)
                logger.info("✅ Plotly Chart rendered to PNG using Kaleido.")
                return buffer
            except Exception as e:
                logger.error(f"Kaleido failed to render Plotly to PNG: {e}")

        logger.error(
            "❌ Failed to render Plotly chart to PNG. Neither WeasyPrint nor Kaleido worked.")
        return None

    def create_category_pie_chart(self, df: pd.DataFrame, title: str) -> Optional[io.BytesIO]:
        if not self.PLOTLY_AVAILABLE or df.empty:
            return None
        category_summary = df.groupby(
            'category')['amount'].sum().sort_values(ascending=False)
        labels = category_summary.index.tolist()
        values = category_summary.values.tolist()
        total = sum(values)
        if total == 0:
            return None

        fig = go.Figure(data=[go.Pie(
            labels=labels,
            values=values,
            hole=.6,
            textinfo='label+percent',
            hoverinfo='label+value+percent',
            insidetextorientation='radial',
            marker_colors=px_colors.qualitative.Pastel1
        )])

        fig.update_layout(
            title_text=title,
            title_font_family=self.myanmar_font_family,
            font_family=self.myanmar_font_family,
            legend_title_text='အမျိုးအစားများ',
            legend_font_family=self.myanmar_font_family,
            margin=dict(l=20, r=20, t=50, b=20),
            showlegend=True
        )
        return self._render_plotly_fig_to_png(fig, width=600, height=600)

    def create_category_bar_chart(self, df: pd.DataFrame, title: str) -> Optional[io.BytesIO]:
        if not self.PLOTLY_AVAILABLE or df.empty:
            return None

        category_summary = df.groupby(
            'category')['amount'].sum().sort_values(ascending=True)
        labels = category_summary.index.tolist()
        values = category_summary.values.tolist()

        fig = go.Figure(data=[go.Bar(
            y=labels,
            x=values,
            orientation='h',
            text=[f'{v:,.0f} Ks' for v in values],
            textposition='outside',
            marker_color=px_colors.qualitative.Pastel1[:len(labels)]
        )])

        fig.update_layout(
            title_text=title,
            title_font_family=self.myanmar_font_family,
            font_family=self.myanmar_font_family,
            xaxis_title="ပမာဏ (Ks)",
            xaxis_title_font_family=self.myanmar_font_family,
            yaxis_title="",
            yaxis_tickfont_family=self.myanmar_font_family,
            margin=dict(l=150, r=50, t=80, b=50),
            height=max(400, len(labels) * 50)
        )
        return self._render_plotly_fig_to_png(fig, width=900, height=max(400, len(labels) * 50))

# ====================================================================
# MyanmarFinanceBot Class (UPDATED for Step 4)
# ====================================================================


class MyanmarFinanceBot:

    def __init__(self):
        self.data_manager = DatabaseManager()
        self.export_manager = ExportManager(EXPORT_DIR)
        self.chart_manager = PlotlyChartManager()
        self.scheduler = AsyncIOScheduler()

        try:
            # Render Environment ကနေ ADMIN_ID ကို ဖတ်ပါ
            self.ADMIN_ID = int(os.getenv('ADMIN_ID')) 
        except (TypeError, ValueError):
            print("❌ CRITICAL: ADMIN_ID environment variable is not set or invalid.")
            self.ADMIN_ID = 0 # Failsafe
        
        self.application: Optional[Application] = None

    # --- Utility: Premium Check ---
    async def check_premium(self, user_id: int, context: ContextTypes.DEFAULT_TYPE) -> bool:
        status = self.data_manager.get_premium_status(user_id)
        if status['is_premium']:
            return True

        keyboard = [
            [InlineKeyboardButton("⭐️ Premium Plan ယူရန်",
                                  callback_data='premium_0')],
            [InlineKeyboardButton("🎁 ၇ ရက် Free Trial ယူရန်",
                                  callback_data='premium_1')]
        ]

        await context.bot.send_message(
            user_id,
            TEXTS["premium_paywall"],
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return False

    # --- Handler: Goal Tracking Menu ---
    async def goal_tracking_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        if not await self.check_premium(user_id, context):
            return

        goals = self.data_manager.get_all_goals(user_id)

        keyboard = [
            [InlineKeyboardButton("➕ ပန်းတိုင်အသစ် ထည့်ရန်", callback_data='goal_add_start')],
            [InlineKeyboardButton("👀 ပန်းတိုင်များ အခြေအနေကြည့်ရန်", callback_data='goal_view_progress')],
            [InlineKeyboardButton(TEXTS["info_button_text"], callback_data='info_goal_tracking')] # <-- ထည့်ရန်
        ]

        if goals:
            keyboard.append([InlineKeyboardButton(
                "🗑️ ပန်းတိုင် ဖျက်ရန်", callback_data='goal_delete_menu')])

        message_text = TEXTS['goal_menu_header']

        if update.callback_query:
            await update.callback_query.edit_message_text(message_text, parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(keyboard))
        else:
            await context.bot.send_message(user_id, message_text, parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(keyboard))

    # --- Handler: View Goal Progress ---
    async def view_goal_progress(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        await update.callback_query.answer("🎯 Goal Progress ကို စစ်ဆေးနေပါသည်...", show_alert=False)

        progress_list = self.data_manager.calculate_goal_progress(user_id)

        if not progress_list:
            await update.callback_query.edit_message_text(TEXTS["goal_no_set"], reply_markup=None)
            return

        details = []
        for p in progress_list:
            details.append(
                TEXTS["goal_progress_detail"].format(
                    emoji=p['emoji'],
                    name=p['name'],
                    amount=p['target_amount'],
                    date=p['target_date'],
                    current_savings=p['current_savings'],
                    remaining=p['remaining'],
                    progress=p['progress']
                )
            )

        message_text = TEXTS["goal_progress_header"] + \
            "\n" + "\n".join(details)

        keyboard = [
            [InlineKeyboardButton("➕ ပန်းတိုင်အသစ် ထည့်ရန်",
                                  callback_data='goal_add_start')],
            [InlineKeyboardButton("🗑️ ပန်းတိုင် ဖျက်ရန်",
                                  callback_data='goal_delete_menu')]
        ]

        await update.callback_query.edit_message_text(message_text, parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(keyboard))

    # --- Handler: Goal Delete Menu ---
    async def delete_goal_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        goals = self.data_manager.get_all_goals(user_id)

        if not goals:
            await update.callback_query.edit_message_text(TEXTS["goal_no_set"], reply_markup=None)
            return

        keyboard = []
        for goal in goals:
            tx_label = f"🗑️ {goal['name']} ({goal['target_amount']:,.0f} Ks)"
            keyboard.append([InlineKeyboardButton(
                tx_label, callback_data=f'goal_delete_confirm_{goal["id"]}')])

        keyboard.append([InlineKeyboardButton(
            "↩️ မီနူးသို့ ပြန်သွားရန်", callback_data='goal_tracking_menu')])

        await update.callback_query.edit_message_text(
            TEXTS['goal_delete_menu'],
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    # --- Handler: Start Add Goal Flow ---
    async def start_add_goal_flow(self, user_id: int, context: ContextTypes.DEFAULT_TYPE):
        context.user_data['mode'] = 'add_goal'

        await context.bot.send_message(
            user_id,
            TEXTS["goal_add_prompt"],
            parse_mode=ParseMode.MARKDOWN
        )

    # --- Handler: Handle Goal Input ---
    async def handle_goal_input(self, update: Update, context: ContextTypes.DEFAULT_TYPE, user_id: int, text: str):
        state = context.user_data

        if state.get('mode') != 'add_goal':
            return False

        parts = text.split(maxsplit=2)

        if len(parts) < 3:
            await update.message.reply_text(TEXTS["goal_invalid_format"])
            return True

        goal_name = parts[0].strip()
        try:
            goal_amount = int(parts[1].replace(',', '').replace('.', ''))
        except ValueError:
            await update.message.reply_text(TEXTS["goal_invalid_format"])
            return True

        target_date_str = parts[2].strip()
        target_date = self._parse_date(target_date_str)

        if target_date is None or target_date.date() < dt.datetime.now().date():
            await update.message.reply_text("❌ ရည်မှန်းချက်ထားသည့် ရက်စွဲပုံစံ မမှန်ကန်ပါ သို့မဟုတ် လက်ရှိရက်စွဲထက် နောက်မကျရပါ။ `MM/DD/YYYY` ပုံစံဖြင့်သာ ရိုက်ထည့်ပေးပါ။")
            return True

        self.data_manager.add_goal(
            user_id, goal_name, goal_amount, target_date)

        await update.message.reply_text(
            TEXTS["goal_add_success"].format(
                name=goal_name,
                amount=goal_amount,
                date=target_date.strftime('%Y-%m-%d')
            )
        )
        context.user_data.clear()
        return True

    # --- Handler: Analytics Menu ---
    async def analytics_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        if not await self.check_premium(user_id, context):
            return

        today = dt.datetime.now()
        month_str = format_myanmar_date(today)

        keyboard = [
            [InlineKeyboardButton(
                f"📈 {month_str} ထွက်ငွေ (Pie Chart)", callback_data='analytics_expense_monthly_pie')],
            [InlineKeyboardButton(
                f"📉 {month_str} ထွက်ငွေ (Bar Chart)", callback_data='analytics_expense_monthly_bar')],
            [InlineKeyboardButton(
                f"💰 {month_str} ဝင်ငွေ (Bar Chart)", callback_data='analytics_income_monthly_bar')],
            [InlineKeyboardButton(
                TEXTS["info_button_text"], callback_data='info_analytics')] # <-- ထည့်ရန်
        ]

        message_text = f"📊 **အသေးစိတ် ခွဲခြမ်းစိတ်ဖြာချက်**\n\nသင်လိုချင်သော ခွဲခြမ်းစိတ်ဖြာချက် အမျိုးအစားကို ရွေးချယ်နိုင်ပါသည်။"
        if update.callback_query:
            await update.callback_query.edit_message_text(message_text, parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(keyboard))
        else:
            await context.bot.send_message(user_id, message_text, parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(keyboard))

    # --- Handler: Custom Category Menu ---

    async def custom_category_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        if not await self.check_premium(user_id, context):
            return

        keyboard = [
            [InlineKeyboardButton("➕ Category အသစ်ထည့်ရန်", callback_data='cat_add')],
            [InlineKeyboardButton("➖ Category ဖယ်ရှားရန်", callback_data='cat_remove')],
            [InlineKeyboardButton(TEXTS["info_button_text"], callback_data='info_custom_category')] # <-- ထည့်ရန်
        ]

        if update.callback_query:
            await update.callback_query.edit_message_text(TEXTS['custom_category_menu'], parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(keyboard))
        else:
            await context.bot.send_message(user_id, TEXTS['custom_category_menu'], parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(keyboard))

    # --- Core Handlers (start, help, summary, budget_status, monthly_report, add_income, add_expense, grant_premium_command) ---
    async def start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user = update.effective_user

        # (!!!) အမှားပြင်ဆင်ချက်- User က /start နှိပ်တိုင်း state ကို အမြဲတမ်း clear လုပ်ပါ (!!!)
        context.user_data.clear()

        self.data_manager.get_premium_status(user.id)

        keyboard = [[KeyboardButton(text) for text in row]
                    for row in TEXTS["main_reply_buttons"]]
        reply_markup = ReplyKeyboardMarkup(
            keyboard, resize_keyboard=True, one_time_keyboard=False)
        await update.message.reply_html(f"မင်္ဂလာပါ {user.mention_html()}! {TEXTS['start_welcome']}", reply_markup=reply_markup)

    async def help(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        await update.message.reply_text(TEXTS["help_text"], parse_mode=ParseMode.MARKDOWN)

    async def privacy(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        await update.message.reply_text(TEXTS["privacy_policy"], parse_mode=ParseMode.MARKDOWN)

    async def delete_my_data_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id

        keyboard = [
            [InlineKeyboardButton(
                TEXTS["delete_data_final_confirm"], callback_data='delete_my_data_confirm')],
            [InlineKeyboardButton(
                TEXTS["delete_data_cancel"], callback_data='delete_my_data_cancel')]
        ]

        await context.bot.send_message(
            user_id,
            TEXTS["delete_data_confirm"],
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    async def premium_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        status = self.data_manager.get_premium_status(user_id)

        if status['is_premium']:
            end_date = dt.datetime.strptime(
                status['end_date'], '%Y-%m-%d').strftime('%Y-%m-%d')
            message_text = f"✅ **Premium Active!**\n\n**ကုန်ဆုံးရက်:** `{end_date}`\n\nPremium လုပ်ဆောင်ချက်အားလုံးကို သုံးစွဲနိုင်ပါပြီ။\n\n{TEXTS['premium_menu_content']}"

            premium_buttons = [
                # Row 1: Analytics
                [
                    InlineKeyboardButton("📊 အသေးစိတ် ခွဲခြမ်းစိတ်ဖြာချက်", callback_data='open_analytics_menu'),
                    InlineKeyboardButton(TEXTS["info_button_text"], callback_data='info_analytics')
                ],
                # Row 2: AI Analysis
                [
                    InlineKeyboardButton(TEXTS["ai_analysis_button"], callback_data='ai_analysis'),
                    InlineKeyboardButton(TEXTS["info_button_text"], callback_data='info_ai_analysis')
                ],
                # Row 3: Custom Category
                [
                    InlineKeyboardButton("🏷️ စိတ်ကြိုက် Category", callback_data='open_custom_category_menu'),
                    InlineKeyboardButton(TEXTS["info_button_text"], callback_data='info_custom_category')
                ],
                # Row 4: Custom Report
                [
                    InlineKeyboardButton("📄 Custom Report", callback_data='start_custom_report'),
                    InlineKeyboardButton(TEXTS["info_button_text"], callback_data='info_custom_report')
                ],
                # Row 5: Goal Tracking
                [
                    InlineKeyboardButton("🎯 ငွေကြေး ပန်းတိုင်များ", callback_data='goal_tracking_menu'),
                    InlineKeyboardButton(TEXTS["info_button_text"], callback_data='info_goal_tracking')
                ]
            ]
            reply_markup = InlineKeyboardMarkup(premium_buttons)
        else:
            message_text = TEXTS['premium_menu_header'] + "\n\n" + \
                TEXTS['premium_menu_content'] + \
                "\n\n" + TEXTS['premium_paywall']
            keyboard = [
                [InlineKeyboardButton(
                    "⭐️ Premium Plan ယူရန်", callback_data='premium_0')],
                [InlineKeyboardButton(
                    "🎁 ၇ ရက် Free Trial ယူရန်", callback_data='premium_1')]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)

        await context.bot.send_message(user_id, message_text, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup)

    async def summary(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        today = dt.datetime.now()
        start_of_month = today.replace(day=1, hour=0, minute=0, second=0)

        transactions = self.data_manager.get_transactions(
            user_id, start_date=start_of_month)

        if not transactions:
            await context.bot.send_message(user_id, TEXTS["no_data"])
            return

        df = pd.DataFrame(transactions)
        total_income = df[df['type'] == 'income']['amount'].sum()
        total_expense = df[df['type'] == 'expense']['amount'].sum()
        balance = total_income - total_expense
        month_str = format_myanmar_date(today)
        response_text = f"{TEXTS['summary_current_month'].format(month=month_str)}\n{TEXTS['summary_details'].format(income=total_income, expense=total_expense, balance=balance)}"
        await context.bot.send_message(user_id, response_text)

    # --- Reminder Menu Handler ---
    async def reminder_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        if not await self.check_premium(user_id, context):
            return

        settings = self.data_manager.get_reminder_settings(user_id)

        weekly_status = "✅ ဖွင့်ထားသည်" if settings.get(
            'weekly_summary') else "❌ ပိတ်ထားသည်"
        weekly_day = settings.get('weekly_day', 'Sunday')
        weekly_text = f"အပတ်စဉ် အစီရင်ခံစာ ({weekly_day}): {weekly_status}"

        daily_status = "✅ ဖွင့်ထားသည်" if settings.get(
            'daily_transaction') else "❌ ပိတ်ထားသည်"
        daily_text = f"နေ့စဉ် ငွေစာရင်း သတိပေးချက် (၉ နာရီ/ ည ၇ နာရီ): {daily_status}"

        keyboard = [
            [
                InlineKeyboardButton(daily_text, callback_data='toggle_daily_reminder'),
                InlineKeyboardButton(TEXTS["info_button_text"], callback_data='info_daily_reminder') # <-- ထည့်ရန်
            ],
            [
                InlineKeyboardButton(weekly_text, callback_data='weekly_reminder_select_day'),
                InlineKeyboardButton(TEXTS["info_button_text"], callback_data='info_weekly_summary') # <-- ထည့်ရန်
            ]
        ]

        if update.callback_query:
            await update.callback_query.edit_message_text(TEXTS["reminder_set_start"], parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(keyboard))
        else:
            await context.bot.send_message(user_id, TEXTS["reminder_set_start"], parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(keyboard))

    # --- Budget Status Calculation Logic ---

    def calculate_budget_status(self, user_id: int, current_tx: Optional[Dict[str, Any]] = None) -> Tuple[str, List[Dict[str, Any]], bool]:
        budgets = self.data_manager.get_budgets(user_id)
        if not budgets:
            return TEXTS["budget_no_set"], [], False

        today = dt.datetime.now()
        month_str = format_myanmar_date(today)
        start_of_month = today.replace(day=1, hour=0, minute=0, second=0)

        transactions = self.data_manager.get_transactions(
            user_id, start_date=start_of_month)

        if not transactions:
            df = pd.DataFrame(columns=['type', 'category', 'amount'])
        else:
            df = pd.DataFrame(transactions)

        expense_df = df[df['type'] == 'expense']

        if today.month == 12:
            end_of_month = dt.datetime(
                today.year + 1, 1, 1) - dt.timedelta(days=1)
        else:
            end_of_month = dt.datetime(
                today.year, today.month + 1, 1) - dt.timedelta(days=1)

        days_remaining = (end_of_month.date() - today.date()).days + 1

        budget_details = []
        alert_needed = False

        for category, budgeted_amount in budgets.items():
            spent = expense_df[expense_df['category']
                               == category]['amount'].sum()
            remaining = budgeted_amount - spent
            percent_spent = (spent / budgeted_amount) * \
                100 if budgeted_amount > 0 else 0

            if (current_tx and
                current_tx.get('category') == category and
                percent_spent >= 80 and
                    percent_spent - ((current_tx['amount'] / budgeted_amount) * 100) < 80):
                alert_needed = True

            status_emoji = "✅"
            if spent > budgeted_amount:
                status_emoji = "❌"
            elif remaining < (budgeted_amount * 0.1):
                status_emoji = "⚠️"

            daily_limit_text = ""
            if days_remaining > 0:
                daily_limit = max(0, remaining) / days_remaining
                daily_limit_text = TEXTS["budget_daily_spending"].format(
                    days_remaining=days_remaining,
                    daily_limit=daily_limit
                )

            budget_details.append(
                f"{status_emoji} **{category}**\n"
                f"  - သတ်မှတ်ဘတ်ဂျက်: {budgeted_amount:,.0f} Ks\n"
                f"  - သုံးစွဲပြီး: {spent:,.0f} Ks ({percent_spent:.1f}%)\n"
                f"  - ကျန်ရှိငွေ: {remaining:,.0f} Ks"
                f"{daily_limit_text}"
            )

        response_text = TEXTS["budget_status_details"].format(
            month=month_str,
            details='\n\n'.join(budget_details)
        )

        return response_text, budget_details, alert_needed

    # --- Budget Status Handler ---
    async def budget_status(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id

        if not self.data_manager.get_premium_status(user_id)['is_premium']:
            await self.check_premium(user_id, context)
            return

        response_text, _, _ = self.calculate_budget_status(user_id)

        if response_text == TEXTS["budget_no_set"]:
            await context.bot.send_message(user_id, response_text)
        else:
            await context.bot.send_message(user_id, response_text, parse_mode=ParseMode.MARKDOWN)

    async def monthly_report(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not self.data_manager.get_premium_status(update.effective_user.id)['is_premium']:
            await self.check_premium(update.effective_user.id, context)
            return

        user_id = update.effective_user.id
        all_txs = self.data_manager.get_transactions(
            user_id, start_date=None, end_date=None)
        if not all_txs:
            await context.bot.send_message(user_id, TEXTS["no_data"])
            return

        dates = []
        for tx in all_txs:
            try:
                dates.append(dt.datetime.fromisoformat(tx['date']))
            except (ValueError, TypeError):
                pass

        unique_months = sorted(list(set((d.year, d.month)
                               for d in dates)), reverse=True)
        keyboard = [[InlineKeyboardButton(format_myanmar_date(dt.datetime(
            year, month, 1)), callback_data=f'select_month_{year}-{month}')] for year, month in unique_months]
        
        # (!!!) အောက်က (၂) ခုကို အသစ် ပြင်ဆင်/ထည့်သွင်းပါ (!!!)
        keyboard.append([
            InlineKeyboardButton("📄 Custom Date Report", callback_data='start_custom_report'),
            InlineKeyboardButton(TEXTS["info_button_text"], callback_data='info_custom_report') # <-- Info Button ထည့်ရန်
        ])
        
        keyboard.append([InlineKeyboardButton(
            TEXTS["info_button_text"], callback_data='info_monthly_report')]) # <-- Info Button ထည့်ရန်

        await context.bot.send_message(user_id, TEXTS["export_select_month"], reply_markup=InlineKeyboardMarkup(keyboard))

    async def add_income(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        await update.message.reply_text(TEXTS["start_add_income"], parse_mode=ParseMode.MARKDOWN)

    async def add_expense(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        await update.message.reply_text(TEXTS["start_add_expense"], parse_mode=ParseMode.MARKDOWN)

    async def grant_premium_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        if user_id != self.ADMIN_ID:
            await update.message.reply_text("🚫 သင့်မှာ Admin သုံးစွဲခွင့်မရှိပါ။")
            return

        if len(context.args) < 2:
            await update.message.reply_text("အသုံးပြုပုံ: /grant_premium [User_ID] [ရက်အရေအတွက်]\nဥပမာ။ ။ `/grant_premium 123456789 30`")
            return

        try:
            target_user_id = int(context.args[0])
            days = int(context.args[1])
        except ValueError:
            await update.message.reply_text("User ID နှင့် ရက်အရေအတွက်ကို ဂဏန်းဖြင့်သာ ထည့်ပါ။")
            return

        end_date = self.data_manager.grant_premium(target_user_id, days)
        await update.message.reply_text(f"✅ User ID: {target_user_id} ကို Premium {days} ရက် ( {end_date} ) အထိ ဖွင့်ပေးလိုက်ပါပြီ။")

        try:
            await context.bot.send_message(
                target_user_id,
                TEXTS["premium_granted"].format(end_date=end_date),
                parse_mode=ParseMode.MARKDOWN
            )
        except Exception as e:
            logger.error(
                f"Cannot send notification to user {target_user_id}: {e}")

    # --- Utility for Date Parsing ---
    def _parse_date(self, date_str: str) -> Optional[dt.datetime]:
        """Parses both MM/DD/YYYY and ISO date strings"""
        if not date_str:
            return None
        try:
            # Try MM/DD/YYYY first (from user input)
            return dt.datetime.strptime(date_str.strip(), '%m/%d/%Y')
        except ValueError:
            try:
                # Try ISO format (from database/callbacks)
                return dt.datetime.fromisoformat(date_str.strip())
            except (ValueError, TypeError):
                return None

    # --- Handler: Start Custom Report Flow ---
    async def start_custom_report_flow(self, user_id: int, context: ContextTypes.DEFAULT_TYPE):
        if not await self.check_premium(user_id, context):
            return

        context.user_data.update(
            {'mode': 'custom_report', 'step': 'start_date', 'start_date': None, 'end_date': None})

        await context.bot.send_message(user_id, TEXTS["custom_report_start"], parse_mode=ParseMode.MARKDOWN)
        await context.bot.send_message(user_id, TEXTS["custom_report_prompt_start_date"], parse_mode=ParseMode.MARKDOWN)

    # --- Handler: Handle Date Input ---
    async def handle_custom_date_input(self, update: Update, context: ContextTypes.DEFAULT_TYPE, user_id: int, text: str):
        state = context.user_data

        if state.get('mode') != 'custom_report':
            return False

        parsed_date = self._parse_date(text)

        if parsed_date is None:
            if state.get('step') == 'start_date':
                await update.message.reply_text(TEXTS["custom_report_invalid_date"] + "\n\n" + TEXTS["custom_report_prompt_start_date"], parse_mode=ParseMode.MARKDOWN)
            elif state.get('step') == 'end_date':
                await update.message.reply_text(TEXTS["custom_report_invalid_date"] + "\n\n" + TEXTS["custom_report_prompt_end_date"], parse_mode=ParseMode.MARKDOWN)
            return True

        if state.get('step') == 'start_date':
            context.user_data['start_date'] = parsed_date.isoformat()
            context.user_data['step'] = 'end_date'

            await update.message.reply_text(f"✅ စတင်ရက်စွဲ: `{parsed_date.strftime('%Y-%m-%d')}`", parse_mode=ParseMode.MARKDOWN)
            await update.message.reply_text(TEXTS["custom_report_prompt_end_date"], parse_mode=ParseMode.MARKDOWN)

        elif state.get('step') == 'end_date':
            start_date_iso = state['start_date']
            start_date = dt.datetime.fromisoformat(start_date_iso)

            if parsed_date.date() < start_date.date():
                await update.message.reply_text("❌ ပြီးဆုံးရက်စွဲသည် စတင်ရက်စွဲထက် စော၍ မရပါ။ ကျေးဇူးပြု၍ ပြန်ရိုက်ပါ။", parse_mode=ParseMode.MARKDOWN)
                return True

            context.user_data['end_date'] = parsed_date.isoformat()
            context.user_data['step'] = 'report_type'

            await update.message.reply_text(f"✅ ပြီးဆုံးရက်စွဲ: `{parsed_date.strftime('%Y-%m-%d')}`", parse_mode=ParseMode.MARKDOWN)

            export_options = [
                [InlineKeyboardButton(
                    "PDF (.pdf)", callback_data='export_custom_pdf')],
                [InlineKeyboardButton(
                    "Excel (.xlsx)", callback_data='export_custom_xlsx')]
            ]
            await update.message.reply_text(TEXTS["export_select_type"], reply_markup=InlineKeyboardMarkup(export_options))

        return True

    # --- Handler: Handle Edit Transaction Input ---
    async def handle_edit_input(self, update: Update, context: ContextTypes.DEFAULT_TYPE, user_id: int, text: str):
        state = context.user_data

        if state.get('mode') != 'edit_transaction':
            return False

        parts = text.split(maxsplit=2)
        if len(parts) != 3:
            await update.message.reply_text(TEXTS["invalid_format"])
            return True

        command = parts[0].lower()
        new_type = 'income' if command in ["ဝင်ငွေ", "income"] else 'expense'
        try:
            new_amount = int(parts[1].replace(',', '').replace('.', ''))
        except ValueError:
            await update.message.reply_text(TEXTS["invalid_format"])
            return True

        new_description = parts[2].strip()
        tx_id = state['tx_id']

        all_categories = self.data_manager.get_all_categories(
            user_id, new_type, TEXTS[f"{new_type}_categories"])
        new_category = next(
            (c for c in all_categories if c in new_description), all_categories[-1])

        updated_tx = self.data_manager.update_transaction(
            user_id, tx_id, new_type, new_amount, new_description, new_category)

        if updated_tx:
            await update.message.reply_text(
                TEXTS["tx_edit_success"].format(
                    new_type=new_type.upper(),
                    new_amount=new_amount,
                    new_category=new_category
                )
            )
        else:
            await update.message.reply_text(TEXTS["tx_not_found"])

        context.user_data.clear()
        return True

    # --- Handle Recurring Transaction Input ---
    async def handle_recurring_tx_input(self, update: Update, context: ContextTypes.DEFAULT_TYPE, user_id: int, text: str):
        state = context.user_data

        if state.get('mode') != 'add_recurring_tx':
            return False

        parts = text.split(maxsplit=3)
        if len(parts) != 4:
            await update.message.reply_text(TEXTS["recurring_tx_invalid_format"])
            return True

        command = parts[0].lower()
        tx_type = 'income' if command in ["ဝင်ငွေ", "income"] else 'expense'

        try:
            amount = int(parts[1].replace(',', '').replace('.', ''))
            day = int(parts[3])
        except ValueError:
            await update.message.reply_text(TEXTS["recurring_tx_invalid_format"])
            return True

        if not (1 <= day <= 28):
            await update.message.reply_text(TEXTS["recurring_tx_invalid_format"])
            return True

        description = parts[2].strip()

        all_categories = self.data_manager.get_all_categories(
            user_id, tx_type, TEXTS[f"{tx_type}_categories"])
        category = next(
            (c for c in all_categories if c in description), all_categories[-1])

        self.data_manager.add_recurring_tx(
            user_id, tx_type, amount, description, category, day)

        await update.message.reply_text(
            TEXTS["recurring_tx_add_success"].format(
                desc=description,
                amount=amount,
                day=day
            )
        )

        context.user_data.clear()
        return True

    # --- (STEP 4) NEW: Handle Admin Find User ID Input ---
    async def admin_handle_find_user_id(self, update: Update, context: ContextTypes.DEFAULT_TYPE, user_id: int, text: str):
        state = context.user_data

        if state.get('mode') != 'admin_find_user_id' or user_id != self.ADMIN_ID:
            return False

        if text.lower() == 'cancel':
            await update.message.reply_text("❌ User ရှာဖွေခြင်းကို ပယ်ဖျက်လိုက်ပါသည်။")
            context.user_data.clear()
            return True

        try:
            target_user_id = int(text.strip())
        except ValueError:
            await update.message.reply_text("❌ User ID (ဂဏန်း) ကိုသာ ရိုက်ထည့်ပါ။")
            return True

        # Admin can't find himself (to avoid confusion with edit_message)
        if target_user_id == user_id:
            await update.message.reply_text("ℹ️ Admin Dashboard မှ Admin ကိုယ်တိုင်၏ Data ကို ရှာဖွေ၍ မရပါ။")
            return True

        details = self.data_manager.get_user_details(target_user_id)

        if not details:
            await update.message.reply_text(TEXTS["admin_user_not_found"].format(user_id=target_user_id))
            context.user_data.clear()
            return True

        status_text = "Active ✅" if details['is_premium'] else "Inactive ❌"

        message = TEXTS["admin_user_details"].format(
            id=details['id'],
            status=status_text,
            end_date=details['end_date'],
            tx_count=details['tx_count'],
            used_trial="Yes" if details['used_trial'] else "No"
        )

        keyboard = [
            [InlineKeyboardButton(
                TEXTS["admin_grant_button"], callback_data=f'admin_grant_user_{target_user_id}')],
            [InlineKeyboardButton(TEXTS["admin_revoke_button"],
                                  callback_data=f'admin_revoke_user_{target_user_id}')],
            [InlineKeyboardButton(
                "↩️ Admin Menu သို့ ပြန်သွားရန်", callback_data='admin_dashboard')]
        ]

        await update.message.reply_text(message, parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(keyboard))
        context.user_data.clear()
        return True

    # --- (STEP 4) NEW: Handle Admin Broadcast Message Input ---
    async def admin_handle_broadcast_message(self, update: Update, context: ContextTypes.DEFAULT_TYPE, user_id: int, text: str):
        state = context.user_data

        if state.get('mode') != 'admin_broadcast_message' or user_id != self.ADMIN_ID:
            return False

        if text.lower() == 'cancel':
            await update.message.reply_text(TEXTS["admin_broadcast_cancelled"])
            context.user_data.clear()
            return True

        context.user_data['broadcast_message'] = text
        context.user_data['mode'] = 'admin_broadcast_confirm'  # Change mode

        user_ids = self.data_manager.get_all_user_ids()
        count = len(user_ids)

        keyboard = [
            [InlineKeyboardButton(TEXTS["admin_broadcast_confirm_button"],
                                  callback_data='admin_broadcast_confirm_send')],
            [InlineKeyboardButton(
                TEXTS["admin_broadcast_cancel_button"], callback_data='admin_broadcast_cancel')]
        ]

        await update.message.reply_text(
            TEXTS["admin_broadcast_confirm"].format(count=count, message=text),
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return True

    # --- MAIN MESSAGE HANDLER (UPDATED for Step 6) ---

    async def handle_message(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        if not update.message or not update.message.text:
            return

        text = update.message.text.strip()
        user_id = update.effective_user.id
        user = update.effective_user
        user_state = context.user_data

        # --- (!!!) NEW: Handle Add Account State (!!!) ---
        if user_state.get('mode') == 'add_account':
            if text.lower() == 'cancel':
                await update.message.reply_text("❌ Account ထည့်သွင်းခြင်းကို ပယ်ဖျက်လိုက်ပါသည်။")
                context.user_data.clear()
                # await self.account_menu(update, context) # Go back to account menu
                return

            parts = text.split(maxsplit=1)
            account_name = parts[0].strip()
            initial_balance = 0
            
            if len(parts) > 1:
                try:
                    # User က "Bank 100,000" လို့ ရိုက်တာကို ခွဲထုတ်ပါ
                    initial_balance = int(parts[1].replace(',', '').replace('.', ''))
                except ValueError:
                    await update.message.reply_text("❌ ပမာဏ မမှန်ကန်ပါ။ ဥပမာ- `Bank 100000`")
                    return

            # DB Manager ကို ခေါ်သုံးပါ
            account, message = self.data_manager.add_account(user_id, account_name, initial_balance)
            
            await update.message.reply_text(message) # Show success or fail message
            
            if account: # If successful
                context.user_data.clear()
                # Account menu ကို ပြန်ခေါ်ပေးပါ
                await self.account_menu(update, context) 
            
            return
        # --- (!!!) End of New State (!!!) ---

        # --- Handle: Awaiting Feedback State (Corrected - ONE TIME ONLY) ---
        if user_state.get('mode') == 'awaiting_feedback':
            if text.lower() == 'cancel':
                await update.message.reply_text(TEXTS["feedback_cancel"])
                context.user_data.clear()
                return

            # Admin ဆီ ပို့မယ့် Message ကို ပြင်ဆင်ပါ
            admin_message = f"""{TEXTS["feedback_admin_header"]}
{TEXTS["feedback_admin_details"].format(
    user_name=user.mention_html(),
    user_id=user_id,
    feedback_text=text
)}
"""
            try:
                # Admin ဆီ ပို့ပါ
                await context.bot.send_message(
                    chat_id=self.ADMIN_ID,
                    text=admin_message,
                    parse_mode=ParseMode.HTML
                )
                # User ကို ကျေးဇူးတင်ကြောင်း ပြောပါ
                await update.message.reply_text(TEXTS["feedback_success"])
                
            except Exception as e:
                logger.error(f"Failed to send feedback to admin: {e}")
                await update.message.reply_text(TEXTS["feedback_error"])
            
            # State ကို ရှင်းလင်းပါ
            context.user_data.clear()
            return
        # --- (!!!) End of Feedback State (!!!) ---


        # --- (STEP 5) Quick Add Number Check ---
        # ... (ကျန်တဲ့ code တွေ ဒီအတိုင်း ဆက်ထားပါ)

        # --- (STEP 5) Quick Add Number Check ---
        quick_add_match = re.match(r'^\d+$', text)
        # ... (ဒီနေရာက ကျန်တဲ့ code တွေ ဒီအတိုင်း ထားပါ) ...
        if quick_add_match and not user_state:
            try:
                amount = int(text)
                if amount <= 0:
                    raise ValueError("Amount is zero")

                context.user_data['mode'] = 'quick_add_type'
                context.user_data['quick_add_amount'] = amount

                keyboard = [
                    [InlineKeyboardButton(
                        TEXTS["quick_add_type_income"], callback_data='quick_add_type_income')],
                    [InlineKeyboardButton(
                        TEXTS["quick_add_type_expense"], callback_data='quick_add_type_expense')],
                    [InlineKeyboardButton(
                        TEXTS["info_button_text"], callback_data='info_quick_add')] # <-- ထည့်ရန်
                ]

                await update.message.reply_text(
                    TEXTS["quick_add_prompt_type"].format(amount=amount),
                    parse_mode=ParseMode.MARKDOWN,
                    reply_markup=InlineKeyboardMarkup(keyboard)
                )
                return

            except ValueError:
                pass
        # --- End of Quick Add ---

        # --- (STEP 4) Admin State Handlers ---
        if user_id == self.ADMIN_ID:
            if user_state.get('mode') == 'admin_find_user_id':
                if await self.admin_handle_find_user_id(update, context, user_id, text):
                    return
            if user_state.get('mode') == 'admin_broadcast_message':
                if await self.admin_handle_broadcast_message(update, context, user_id, text):
                    return
            if user_state.get('mode') == 'admin_broadcast_confirm':
                await update.message.reply_text("⚠️ ကျေးဇူးပြု၍ အပေါ်က ခလုတ်ကို နှိပ်၍ အတည်ပြုပါ (သို့) ပယ်ဖျက်ပါ။")
                return

        # --- Regular User State Handlers ---
        if user_state.get('mode') == 'add_goal':
            if await self.handle_goal_input(update, context, user_id, text):
                return

        if user_state.get('mode') == 'edit_transaction':
            if await self.handle_edit_input(update, context, user_id, text):
                return

        if user_state.get('mode') == 'custom_report':
            if await self.handle_custom_date_input(update, context, user_id, text):
                return

        if user_state.get('mode') == 'add_category':
            cat_type = user_state['type']
            cat_name = text.strip()

            default_cats = TEXTS['expense_categories'] if cat_type == 'expense' else TEXTS['income_categories']
            if cat_name in default_cats:
                await update.message.reply_text(TEXTS["cat_add_fail_exists"].format(name=cat_name))
                context.user_data.clear()
                return

            if self.data_manager.add_custom_category(user_id, cat_type, cat_name):
                await update.message.reply_text(TEXTS["cat_add_success"].format(name=cat_name))
            else:
                await update.message.reply_text(TEXTS["cat_add_fail_exists"].format(name=cat_name))

            context.user_data.clear()
            return

        if user_state.get('mode') == 'add_recurring_tx':
            if await self.handle_recurring_tx_input(update, context, user_id, text):
                return

        if user_state.get('mode') == 'awaiting_screenshot':
            await update.message.reply_text(TEXTS["premium_awaiting_screenshot"], parse_mode=ParseMode.MARKDOWN)
            return

        # --- (STEP 6) Backup File State Check ---
        if user_state.get('mode') == 'awaiting_backup_file':
            if text.lower() == 'cancel':
                await update.message.reply_text(TEXTS["restore_cancelled"])
                context.user_data.clear()
            else:
                await update.message.reply_text(TEXTS["restore_prompt"], parse_mode=ParseMode.MARKDOWN)
            return

        # 6. --- Handle Main Menu Buttons and Commands ---

        if text == TEXTS["main_reply_buttons"][1][0]:  # 📊 အခြေအနေ
            await self.summary(update, context)
            return
        elif text == TEXTS["main_reply_buttons"][1][1]:  # 🧾 အစီရင်ခံစာ
            if not self.data_manager.get_premium_status(user_id)['is_premium']:
                await self.check_premium(user_id, context)
                return
            await self.monthly_report(update, context)
            return
        elif text == TEXTS["main_reply_buttons"][0][0]:  # 💰 ဝင်ငွေ
            await self.add_income(update, context)
            return
        elif text == TEXTS["main_reply_buttons"][0][1]:  # 💸 ထွက်ငွေ
            await self.add_expense(update, context)
            return
        elif text == TEXTS["main_reply_buttons"][2][1]:  # 🎯 ဘတ်ဂျက်
            if not self.data_manager.get_premium_status(user_id)['is_premium']:
                await self.check_premium(user_id, context)
                return
            await update.message.reply_text(TEXTS["budget_set_start"], parse_mode=ParseMode.MARKDOWN)
            return
        elif text == TEXTS["main_reply_buttons"][2][0]:  # 🗓️ သတိပေးချက်
            await self.reminder_menu(update, context)
            return
        elif text == TEXTS["main_reply_buttons"][3][1]:  # ⭐️ Premium
            await self.premium_menu(update, context)
            return
        elif text == TEXTS["main_reply_buttons"][3][0]:  # ⚙️ စီမံခန့်ခွဲ
            await self.manage_transactions_menu(update, context)
            return
        elif text == TEXTS["main_reply_buttons"][4][0]:  # 💡 အကြံပြုမည်  <-- (!!!) ဒီ Block အသစ်ကို ထည့်ပါ (!!!)
            context.user_data['mode'] = 'awaiting_feedback'
            await update.message.reply_text(TEXTS["feedback_prompt"])
            return
        elif text == TEXTS["main_reply_buttons"][4][1]:  # 🔒 Privacy
            await self.privacy(update, context)
            return

        # 7. --- Handle Transaction/Budget Input (MODIFIED FOR STEP 2) ---
        parts = text.split(maxsplit=2)
        command = parts[0].lower()

        if command in ["ဝင်ငွေ", "income", "ထွက်ငွေ", "expense"]:
            if len(parts) != 3:
                await update.message.reply_text(TEXTS["invalid_format"])
                return
            
            tx_type = 'income' if command in ["ဝင်ငွေ", "income"] else 'expense'
            try:
                amount = int(parts[1].replace(',', '').replace('.', ''))
            except ValueError:
                await update.message.reply_text(TEXTS["invalid_format"])
                return
            description = parts[2].strip()

            all_categories = self.data_manager.get_all_categories(
                user_id, tx_type, TEXTS[f"{tx_type}_categories"])
            category = next(
                (c for c in all_categories if c in description), all_categories[-1])

            # --- (!!!) NEW LOGIC (!!!) ---
            # Data ကို တိုက်ရိုက် မသိမ်းတော့ဘဲ၊ User State ထဲမှာ ခဏ သိမ်းပါ
            context.user_data.clear() # State အဟောင်း ရှင်းပါ
            context.user_data['mode'] = 'awaiting_account_selection'
            context.user_data['tx_data'] = {
                'type': tx_type,
                'amount': amount,
                'description': description,
                'category': category
            }
            
            # User ကို Account ရွေးခိုင်းမယ့် Helper Function ကို ခေါ်ပါ
            prompt_text = TEXTS["select_account_prompt"].format(
                tx_type=command, 
                amount=amount, 
                desc=description
            )
            await self.prompt_account_selection(update.message, context, user_id, prompt_text)
            # --- (!!!) End of New Logic (!!!) ---

            return # Function ကို ဒီမှာတင် ရပ်လိုက်ပါ

        elif command in ["ဘတ်ဂျက်", "budget"]:
            if not self.data_manager.get_premium_status(user_id)['is_premium']:
                await self.check_premium(user_id, context)
                return

            if len(parts) != 3:
                await update.message.reply_text(TEXTS["budget_set_start"])
                return
            try:
                amount = int(parts[1].replace(',', '').replace('.', ''))
            except ValueError:
                await update.message.reply_text(TEXTS["invalid_format"])
                return
            category = parts[2].strip()

            all_expense_cats = self.data_manager.get_all_categories(
                user_id, 'expense', TEXTS["expense_categories"])
            if category not in all_expense_cats:
                await update.message.reply_text(f"❌ '{category}' ဆိုတဲ့ Category မရှိပါဘူး။ အသုံးပြုနိုင်တဲ့ Category တွေကတော့: {', '.join(all_expense_cats)} ဖြစ်ပါတယ်။")
                return
            self.data_manager.set_budget(user_id, category, amount)
            await update.message.reply_text(TEXTS["budget_set_success"].format(category=category, amount=amount))
            return

        await update.message.reply_text(TEXTS["unknown_command"])

    # --- Handle Screenshot for Premium ---

    async def handle_screenshot(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        state = context.user_data

        if state.get('mode') != 'awaiting_screenshot':
            return

        if not update.message or not update.message.photo:
            await update.message.reply_text(TEXTS["premium_awaiting_screenshot"], parse_mode=ParseMode.MARKDOWN)
            return

        logger.info(f"Received screenshot from user {user_id} for premium.")

        months = state.get('plan_months', '1')
        duration_text = TEXTS.get(f"premium_duration_{months}", f"{months} လ")
        price_text = TEXTS.get(f"premium_price_{months}", "N/A")
        days_map = {'1': 30, '6': 180, '12': 365}
        days = days_map.get(str(months), 30)

        user_name = update.effective_user.mention_html()

        try:
            await context.bot.forward_message(
                chat_id=self.ADMIN_ID,
                from_chat_id=user_id,
                message_id=update.message.message_id
            )

            admin_caption = TEXTS["admin_approval_message"].format(
                user_name=user_name,
                user_id=user_id,
                duration_text=duration_text,
                price_text=price_text
            )

            keyboard = [
                [InlineKeyboardButton(
                    f"✅ Approve {days} Days", callback_data=f'admin_approve_{user_id}_{days}')],
                [InlineKeyboardButton(
                    "❌ Reject", callback_data=f'admin_reject_{user_id}')]
            ]

            await context.bot.send_message(
                self.ADMIN_ID,
                admin_caption,
                parse_mode=ParseMode.HTML,
                reply_markup=InlineKeyboardMarkup(keyboard)
            )

            await update.message.reply_text(TEXTS["premium_screenshot_received"], parse_mode=ParseMode.MARKDOWN)
            context.user_data.clear()

        except Exception as e:
            logger.error(
                f"Failed to forward screenshot to admin {self.ADMIN_ID}: {e}")
            await update.message.reply_text(f"❌ Admin ထံသို့ ပို့ရာတွင် အမှားဖြစ်ပွားပါသည်။ ကျေးဇူးပြု၍ Admin (@adu1010101) ကို တိုက်ရိုက် ဆက်သွယ်ပါ။")

    # --- (STEP 6) NEW HANDLER: Handle Backup File (JSON) ---
    async def handle_backup_file(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        state = context.user_data

        if state.get('mode') != 'awaiting_backup_file':
            # If user sends a random file, ignore it
            return

        if not await self.check_premium(user_id, context):
            return

        if not update.message or not update.message.document:
            await update.message.reply_text(TEXTS["restore_prompt"], parse_mode=ParseMode.MARKDOWN)
            return

        document = update.message.document
        if not document.file_name.endswith('.json'):
            await update.message.reply_text(TEXTS["restore_error_json"])
            return

        try:
            # Download the file
            json_file = await document.get_file()
            file_content_bytes = await json_file.download_as_bytearray()

            # Decode and parse JSON
            file_content_str = file_content_bytes.decode('utf-8')
            backup_data = json.loads(file_content_str)

            # Check if file format is valid (key တွေ အကုန် ပါ, မပါ စစ်ပါ)
            required_keys = ["transactions", "budgets",
                             "goals", "custom_categories", "recurring_txs"]
            if not all(key in backup_data for key in required_keys):
                await update.message.reply_text(TEXTS["restore_error_format"])
                context.user_data.clear()
                return

            # Run the restore process in DatabaseManager
            success = self.data_manager.restore_data_from_backup(
                user_id, backup_data)

            if success:
                await update.message.reply_text(TEXTS["restore_success"])
            else:
                await update.message.reply_text(TEXTS["restore_error_general"])

        except json.JSONDecodeError:
            await update.message.reply_text(TEXTS["restore_error_json"])
        except Exception as e:
            logger.error(f"Error restoring backup for user {user_id}: {e}")
            await update.message.reply_text(TEXTS["restore_error_general"])

        context.user_data.clear()

    # --- MAIN CALLBACK HANDLER (UPDATED for Step 6) ---

    async def handle_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        query = update.callback_query
        user_id = query.from_user.id
        data = query.data

        state = context.user_data

        # (!!!) This is correct (!!!)
        if data.startswith('info_'):
            # ... (info logic) ...
            return # Message ကို edit လုပ်စရာမလိုလို့ ဒီမှာတင် ရပ်လိုက်ပါ
        
        # --- (!!!) NEW: Account Callbacks (Corrected 'if' to 'elif') (!!!) ---
        elif data == 'account_menu':
            await self.account_menu(update, context)
            return
        elif data == 'account_view':
            await self.account_view_balances(update, context)
            return
        elif data == 'account_add_start':
            await self.account_add_prompt(update, context)
            return
        # --- (!!!) End of New Callbacks (!!!) ---

        # --- (!!!) NEW: Handle Account Selection for Transaction (FINAL FIX) (!!!) ---
        elif data.startswith('tx_select_account_') and state.get('mode') == 'awaiting_account_selection':
            tx_data = state.get('tx_data')
            
            if not tx_data:
                await query.edit_message_text("❌ အချိန် ကျော်လွန်သွားပါသဖြင့်၊ ကျေးဇူးပြု၍ ငွေစာရင်းကို အစမှ ပြန်ထည့်ပါ။")
                context.user_data.clear()
                return

            account_id = data.replace('tx_select_account_', '')
            
            if account_id == 'none':
                # User က "Account မသတ်မှတ်" ကို ရွေးသည်
                self.data_manager.add_transaction(
                    user_id=user_id,
                    type=tx_data['type'],
                    amount=tx_data['amount'],
                    description=tx_data['description'],
                    category=tx_data['category'],
                    account_id=None # <-- Account ID = None
                )
                await query.edit_message_text(TEXTS["data_saved_no_account"].format(
                    category=tx_data['category'], amount=tx_data['amount']
                ))
            
            else:
                # User က Account တစ်ခုခုကို ရွေးသည်
                self.data_manager.add_transaction(
                    user_id=user_id,
                    type=tx_data['type'],
                    amount=tx_data['amount'],
                    description=tx_data['description'],
                    category=tx_data['category'],
                    account_id=account_id # <-- Account ID အစစ်
                )
                
                # (!!!) --- FINAL FIXED BLOCK (!!!) ---
                # Account နာမည်ကို ပြန်ရှာပြီး user ကို ပြပါ
                account_list = self.data_manager.get_accounts(user_id)
                # 'account' ဆိုတဲ့ variable ထဲကို ရှာတွေ့တဲ့ dict ကို ထည့်ပါ
                account = next((acc for acc in account_list if acc['id'] == account_id), None)
                # 'acc' ကို မသုံးဘဲ၊ 'account' variable ကို သုံးပါ
                account_name = account['name'] if account else "Unknown" 
                # (!!!) --- END OF FIX --- (!!!)
                
                await query.edit_message_text(TEXTS["data_saved_with_account"].format(
                    category=tx_data['category'], 
                    amount=tx_data['amount'], 
                    account_name=account_name
                ))

            context.user_data.clear() # State ကို ရှင်းပါ
            
            # --- (!!!) Real-time Budget Alert Check (ဒီနေရာကို ရွှေ့ပါ) (!!!) ---
            if tx_data['type'] == 'expense' and self.data_manager.get_premium_status(user_id)['is_premium']:
                await self.check_budget_alert(user_id, tx_data['category'], tx_data['amount'], context)
            
            return
        # --- (!!!) End of New Transaction Callback (!!!) ---
        
        # --- (STEP 6) NEW: Backup/Restore Callbacks (Corrected - ONE TIME ONLY) ---
        elif data == 'backup_restore_menu':
            if not await self.check_premium(user_id, context):
                return

            keyboard = [
                [InlineKeyboardButton(
                    TEXTS["backup_button"], callback_data='backup_start')],
                [InlineKeyboardButton(
                    TEXTS["restore_button"], callback_data='restore_start')]
            ]
            await query.edit_message_text(
                TEXTS["backup_restore_menu_header"],
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            return
        # --- (!!!) End of Duplicate block (!!!) ---

        elif data == 'backup_start':
            # ... (ကျန်တဲ့ code တွေ ဒီအတိုင်း ဆက်ထားပါ)
            if not await self.check_premium(user_id, context):
                return
            
            # ... (ကျန်တဲ့ code တွေ ဒီအတိုင်း ဆက်ထားပါ) ...

            await query.edit_message_text(TEXTS["backup_prompt_sending"])

            try:
                # 1. Get all data from DB
                all_data = self.data_manager.get_all_data_for_backup(user_id)

                # 2. Convert to JSON string (use default=str to handle datetime objects)
                backup_json_str = json.dumps(
                    all_data, indent=4, ensure_ascii=False, default=str)

                # 3. Convert string to BytesIO file
                backup_file = io.BytesIO(backup_json_str.encode('utf-8'))

                # 4. Send file to user
                date_str = dt.datetime.now().strftime('%Y-%m-%d')
                file_name = f"backup_adu_finance_{date_str}.json"

                await context.bot.send_document(
                    chat_id=user_id,
                    document=backup_file,
                    filename=file_name,
                    caption=TEXTS["backup_prompt_success"].format(
                        date=date_str)
                )
                await query.delete_message()  # "Processing..." message ကို ဖျက်ပါ

            except Exception as e:
                logger.error(f"Error during backup for user {user_id}: {e}")
                await query.edit_message_text("❌ Backup ပြုလုပ်ရာတွင် အမှားအယွင်း ဖြစ်ပွားပါသည်။")
            return

        elif data == 'restore_start':
            if not await self.check_premium(user_id, context):
                return

            # Set user state to wait for a .json file
            context.user_data['mode'] = 'awaiting_backup_file'
            await query.edit_message_text(TEXTS["restore_prompt"], parse_mode=ParseMode.MARKDOWN)
            return
        # --- End of Backup/Restore Callbacks ---

        # --- (STEP 5.1) NEW: Quick Add Callbacks (REVISED) ---

        # --- (Block 1) Handle Type Selection (ဝင်ငွေ or ထွက်ငွေ) ---
        if data.startswith('quick_add_type_') and state.get('mode') == 'quick_add_type':
            amount = state.get('quick_add_amount')
            if not amount:
                await query.edit_message_text("❌ အချိန် ကျော်လွန်သွားပါသဖြင့်၊ ကျေးဇူးပြု၍ ဂဏန်းကို အသစ် ပြန်ရိုက်ထည့်ပါ။")
                context.user_data.clear()
                return

            transaction_type = ""
            if data == 'quick_add_type_income':
                transaction_type = 'income'
            elif data == 'quick_add_type_expense':
                transaction_type = 'expense'
            else:
                return  # Should not happen

            # --- (!!!) NEW LOGIC (!!!) ---
            # User ကို Category ဆက်ရွေးခိုင်းပါ (ဝင်ငွေ ဖြစ်စေ၊ ထွက်ငွေ ဖြစ်စေ)
            context.user_data['mode'] = 'quick_add_category'
            # <-- (!!!) Save the type
            context.user_data['quick_add_type'] = transaction_type

            all_categories = self.data_manager.get_all_categories(
                user_id, transaction_type, TEXTS[f"{transaction_type}_categories"])

            context.user_data['quick_add_categories'] = all_categories

            keyboard = []
            row = []
            for index, cat in enumerate(all_categories):
                # Use index for callback_data to avoid invalid characters like '/'
                row.append(InlineKeyboardButton(
                    cat, callback_data=f'quick_add_category_{index}'))
                if len(row) == 3:
                    keyboard.append(row)
                    row = []
            if row:
                keyboard.append(row)

            # Change the prompt text based on type
            prompt_text = TEXTS["quick_add_prompt_category"].format(
                amount=amount)  # Default for expense
            if transaction_type == 'income':
                prompt_text = f"💰 **{amount:,.0f} Ks** ကို ဝင်ငွေ (Income) အဖြစ် မှတ်သားပါမည်။\n\n👇 ကျေးဇူးပြု၍ Category တစ်ခု ရွေးချယ်ပါ။"

            await query.edit_message_text(
                prompt_text,
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            return

        # --- (Block 2) Handle Category Selection (MODIFIED FOR STEP 2) ---
        elif data.startswith('quick_add_category_') and state.get('mode') == 'quick_add_category':
            amount = state.get('quick_add_amount')
            all_categories = state.get('quick_add_categories')
            transaction_type = state.get('quick_add_type')

            if not amount or not all_categories or not transaction_type:
                await query.edit_message_text("❌ အချိန် ကျော်လွန်သွားပါသဖြင့်၊ ကျေးဇူးပြု၍ ဂဏန်းကို အသစ် ပြန်ရိုက်ထည့်ပါ။")
                context.user_data.clear()
                return

            try:
                category_index = int(data.replace('quick_add_category_', ''))
                category = all_categories[category_index]
            except (ValueError, IndexError):
                await query.edit_message_text("❌ Category ရွေးချယ်မှု မှားယွင်းပါသည်။")
                context.user_data.clear()
                return

            description = f"Quick Add {category}"

            # --- (!!!) NEW LOGIC (!!!) ---
            # Data ကို တိုက်ရိုက် မသိမ်းတော့ဘဲ၊ User State ထဲမှာ ခဏ သိမ်းပါ
            context.user_data['mode'] = 'awaiting_account_selection'
            context.user_data['tx_data'] = {
                'type': transaction_type,
                'amount': amount,
                'description': description,
                'category': category
            }
            
            # User ကို Account ရွေးခိုင်းမယ့် Helper Function ကို ခေါ်ပါ
            prompt_text = TEXTS["select_account_prompt"].format(
                tx_type=transaction_type, 
                amount=amount, 
                desc=description
            )
            # (!!!) Message ကို Edit လုပ်ပါ (!!!)
            await self.prompt_account_selection(update, context, user_id, prompt_text)
            # --- (!!!) End of New Logic (!!!) ---

            return
        # --- End of Quick Add Callbacks ---

        # --- (STEP 5.1) NEW: AI Analyst Callback ---
        elif data == 'ai_analysis':
            if not await self.check_premium(user_id, context):
                return

            await query.edit_message_text(TEXTS["processing"])

            await self.run_ai_analysis(update, context)
            return
        # --- End of AI Analyst Callback ---

        # --- (STEP 4) NEW: Admin Dashboard Callbacks ---
        if data == 'admin_dashboard':
            if user_id != self.ADMIN_ID:
                return await query.answer(TEXTS["not_admin"], show_alert=True)
            await self.admin_dashboard(update, context)
            return

        if data == 'admin_stats':
            if user_id != self.ADMIN_ID:
                return await query.answer(TEXTS["not_admin"], show_alert=True)
            await self.admin_stats(update, context)
            return

        if data == 'admin_broadcast_prompt':
            if user_id != self.ADMIN_ID:
                return await query.answer(TEXTS["not_admin"], show_alert=True)
            await self.admin_broadcast_prompt(update, context)
            return

        if data == 'admin_find_user_prompt':
            if user_id != self.ADMIN_ID:
                return await query.answer(TEXTS["not_admin"], show_alert=True)
            await self.admin_find_user_prompt(update, context)
            return

        if data == 'admin_broadcast_confirm_send':
            if user_id != self.ADMIN_ID:
                return await query.answer(TEXTS["not_admin"], show_alert=True)
            await self.admin_broadcast_send(update, context)
            return

        if data == 'admin_broadcast_cancel':
            if user_id != self.ADMIN_ID:
                return
            context.user_data.clear()
            await query.edit_message_text(TEXTS["admin_broadcast_cancelled"])
            return

        if data.startswith('admin_grant_user_'):
            if user_id != self.ADMIN_ID:
                return await query.answer(TEXTS["not_admin"], show_alert=True)
            target_user_id = int(data.split('_')[-1])
            self.data_manager.grant_premium(target_user_id, 30)
            await query.answer(TEXTS["admin_user_granted"].format(user_id=target_user_id))
            await self._send_user_details(query.message, context, target_user_id)
            return

        if data.startswith('admin_revoke_user_'):
            if user_id != self.ADMIN_ID:
                return await query.answer(TEXTS["not_admin"], show_alert=True)
            target_user_id = int(data.split('_')[-1])
            self.data_manager.revoke_premium(target_user_id)
            await query.answer(TEXTS["admin_user_revoked"].format(user_id=target_user_id))
            await self._send_user_details(query.message, context, target_user_id)
            return

        # --- (STEP 3) Admin Approval Callbacks ---
        if data.startswith('admin_approve_'):
            if user_id != self.ADMIN_ID:
                return await query.answer(TEXTS["not_admin"], show_alert=True)

            try:
                parts = data.split('_')
                target_user_id = int(parts[2])
                days = int(parts[3])

                end_date = self.data_manager.grant_premium(
                    target_user_id, days)

                await context.bot.send_message(
                    target_user_id,
                    TEXTS["premium_granted"].format(end_date=end_date),
                    parse_mode=ParseMode.MARKDOWN
                )

                await query.edit_message_text(
                    f"✅ Approved!\nUser {target_user_id} ကို {days} ရက် Premium ဖွင့်ပေးလိုက်ပါပြီ။",
                    reply_markup=None
                )

            except Exception as e:
                logger.error(f"Error in admin approval: {e}")
                await query.edit_message_text(f"❌ Approval Error: {e}")
            return

        if data.startswith('admin_reject_'):
            if user_id != self.ADMIN_ID:
                return await query.answer(TEXTS["not_admin"], show_alert=True)

            try:
                parts = data.split('_')
                target_user_id = int(parts[2])

                await context.bot.send_message(
                    target_user_id,
                    TEXTS["user_approval_rejected"],
                    parse_mode=ParseMode.MARKDOWN
                )

                await query.edit_message_text(
                    f"❌ Rejected!\nUser {target_user_id} ကို Premium ပယ်ချကြောင်း အကြောင်းကြားလိုက်ပါပြီ။",
                    reply_markup=None
                )

            except Exception as e:
                logger.error(f"Error in admin rejection: {e}")
                await query.edit_message_text(f"❌ Rejection Error: {e}")
            return

        # --- NEW: Delete Data Callbacks ---
        if data == 'delete_my_data_confirm':
            if self.data_manager.delete_user_data(user_id):
                await query.edit_message_text(TEXTS["delete_data_success"], reply_markup=None)
                logger.info(f"User {user_id} has deleted all their data.")
            else:
                await query.edit_message_text("❌ Data ဖျက်ရန် ရှာမတွေ့ပါ။", reply_markup=None)
            return
        elif data == 'delete_my_data_cancel':
            await query.edit_message_text(TEXTS["delete_data_cancelled"], reply_markup=None)
            return

        # --- Goal Tracking Callbacks ---
        if data == 'goal_tracking_menu':
            await self.goal_tracking_menu(update, context)
            return
        elif data == 'goal_add_start':
            await query.edit_message_text(TEXTS['goal_menu_header'], reply_markup=None)
            await self.start_add_goal_flow(user_id, context)
            return
        elif data == 'goal_view_progress':
            await self.view_goal_progress(update, context)
            return
        elif data == 'goal_delete_menu':
            await self.delete_goal_menu(update, context)
            return
        elif data.startswith('goal_delete_confirm_'):
            goal_id = data.split('_')[3]
            goals = self.data_manager.get_all_goals(user_id)
            goal = next((g for g in goals if g['id'] == goal_id), None)

            if goal and self.data_manager.delete_goal(user_id, goal_id):
                await query.edit_message_text(TEXTS["goal_delete_success"].format(name=goal['name']), reply_markup=None)
            else:
                await query.edit_message_text(TEXTS["goal_not_found"], reply_markup=None)
            return

        # --- Transaction Management Callbacks ---
        elif data.startswith('tx_select_'):
            tx_id = data.split('_')[2]
            tx = self.data_manager.get_transaction_by_id(user_id, tx_id)

            if not tx:
                await query.edit_message_text(TEXTS["tx_not_found"], reply_markup=None)
                return

            tx_date_obj = self._parse_date(tx['date'])
            tx_date = tx_date_obj.strftime(
                '%Y-%m-%d') if tx_date_obj else "N/A"

            tx_type = "ဝင်ငွေ" if tx['type'] == 'income' else "ထွက်ငွေ"
            tx_text = f"⚙️ စီမံခန့်ခွဲနေသော မှတ်တမ်း:\n\n`{tx_date} - {tx_type} ({tx['category']}) : {tx['amount']:,.0f} Ks`\n**ဖော်ပြချက်:** `{tx['description']}`"

            keyboard = [
                [InlineKeyboardButton(
                    "✏️ ပြင်ဆင်ရန်", callback_data=f'tx_edit_start_{tx_id}')],
                [InlineKeyboardButton(
                    "🗑️ ဖျက်ပစ်ရန်", callback_data=f'tx_delete_confirm_{tx_id}')],
                [InlineKeyboardButton(
                    "↩️ မီနူးသို့ ပြန်သွားရန်", callback_data='manage_tx_menu_back')]
            ]

            await query.edit_message_text(tx_text, parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(keyboard))
            return

        elif data.startswith('tx_delete_confirm_'):
            tx_id = data.split('_')[3]
            if self.data_manager.delete_transaction(user_id, tx_id):
                await query.edit_message_text(TEXTS["tx_delete_success"], reply_markup=None)
            else:
                await query.edit_message_text(TEXTS["tx_not_found"], reply_markup=None)
            return

        elif data.startswith('tx_edit_start_'):
            tx_id = data.split('_')[3]
            tx = self.data_manager.get_transaction_by_id(user_id, tx_id)

            if not tx:
                await query.edit_message_text(TEXTS["tx_not_found"], reply_markup=None)
                return

            context.user_data.update(
                {'mode': 'edit_transaction', 'tx_id': tx_id})

            tx_date_obj = self._parse_date(tx['date'])
            tx_date = tx_date_obj.strftime(
                '%Y-%m-%d') if tx_date_obj else "N/A"

            tx_type = "ဝင်ငွေ" if tx['type'] == 'income' else "ထွက်ငွေ"

            await query.edit_message_text(
                TEXTS["tx_edit_prompt"].format(
                    date=tx_date,
                    type=tx_type,
                    category=tx['category'],
                    amount=tx['amount']
                ),
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=None
            )
            return

        elif data == 'manage_tx_menu_back':
            await self.manage_transactions_menu(update, context)
            return

        # --- Recurring Transaction Callbacks ---
        elif data == 'recurring_tx_menu':
            if not await self.check_premium(user_id, context):
                return

            keyboard = [
                [InlineKeyboardButton(
                    "➕ အသစ်ထည့်ရန်", callback_data='recurring_tx_add_start')],
                [InlineKeyboardButton(
                    "🗑️ ဖျက်ပစ်ရန်", callback_data='recurring_tx_delete_menu')],
                [InlineKeyboardButton(
                    "↩️ မီနူးသို့ ပြန်သွားရန်", callback_data='manage_tx_menu_back')]
            ]
            await query.edit_message_text(TEXTS["recurring_tx_menu_header"], parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(keyboard))
            return

        elif data == 'recurring_tx_add_start':
            if not await self.check_premium(user_id, context):
                return
            context.user_data['mode'] = 'add_recurring_tx'
            await query.edit_message_text(TEXTS["recurring_tx_add_prompt"], parse_mode=ParseMode.MARKDOWN, reply_markup=None)
            return

        elif data == 'recurring_tx_delete_menu':
            if not await self.check_premium(user_id, context):
                return

            rtxs = self.data_manager.get_recurring_txs(user_id)
            if not rtxs:
                await query.edit_message_text(TEXTS["recurring_tx_no_set"], reply_markup=None)
                return

            keyboard = []
            for rtx in rtxs:
                tx_type_my = "ဝင်ငွေ" if rtx['type'] == 'income' else "ထွက်ငွေ"
                label = f"🗑️ {rtx['description']} ({tx_type_my}) - လစဉ် {rtx['day']} ရက်နေ့"
                keyboard.append([InlineKeyboardButton(
                    label, callback_data=f'recurring_tx_delete_confirm_{rtx["id"]}')])

            keyboard.append([InlineKeyboardButton(
                "↩️ မီနူးသို့ ပြန်သွားရန်", callback_data='recurring_tx_menu')])
            await query.edit_message_text(TEXTS["recurring_tx_delete_menu"], reply_markup=InlineKeyboardMarkup(keyboard))
            return

        elif data.startswith('recurring_tx_delete_confirm_'):
            if not await self.check_premium(user_id, context):
                return

            tx_id = data.split('_')[4]
            rtx_list = self.data_manager.get_recurring_txs(user_id)
            rtx = next((tx for tx in rtx_list if tx['id'] == tx_id), None)

            if rtx and self.data_manager.delete_recurring_tx(user_id, tx_id):
                await query.edit_message_text(TEXTS["recurring_tx_delete_success"].format(name=rtx['description']), reply_markup=None)
            else:
                await query.edit_message_text(TEXTS["recurring_tx_not_found"], reply_markup=None)
            return

        # --- Premium Feature Placeholders & Menus ---
        elif data == 'start_custom_report':
            await query.edit_message_text("📄 Custom Date Report ကို စတင်ပါမည်။")
            await self.start_custom_report_flow(user_id, context)
            return
        elif data == 'open_analytics_menu':
            await self.analytics_menu(update, context)
            return
        elif data == 'open_custom_category_menu':
            await self.custom_category_menu(update, context)
            return

        # --- Custom Category Callbacks ---
        elif data == 'cat_add':
            keyboard = [[InlineKeyboardButton("💰 ဝင်ငွေ", callback_data='cat_add_income')], [
                InlineKeyboardButton("💸 ထွက်ငွေ", callback_data='cat_add_expense')]]
            await query.edit_message_text(TEXTS["select_cat_type"], reply_markup=InlineKeyboardMarkup(keyboard))
            return
        elif data.startswith('cat_add_'):
            cat_type = data.split('_')[-1]
            context.user_data.update(
                {'mode': 'add_category', 'type': cat_type})
            await query.edit_message_text(TEXTS["add_cat_prompt"])
            return
        elif data == 'cat_remove':
            keyboard = [[InlineKeyboardButton("💰 ဝင်ငွေ", callback_data='cat_remove_income')], [
                InlineKeyboardButton("💸 ထွက်ငွေ", callback_data='cat_remove_expense')]]
            await query.edit_message_text(TEXTS["select_cat_type"], reply_markup=InlineKeyboardMarkup(keyboard))
            return
        elif data.startswith('cat_remove_'):
            cat_type = data.split('_')[-1]
            custom_cats = self.data_manager.get_custom_categories(
                user_id, cat_type)
            if not custom_cats:
                await query.edit_message_text("ℹ️ ဖယ်ရှားရန် စိတ်ကြိုက် Category မရှိသေးပါ။")
                return

            keyboard = [[InlineKeyboardButton(
                f"🗑️ {cat}", callback_data=f'cat_remove_final_{cat_type}_{cat}')] for cat in custom_cats]
            await query.edit_message_text(TEXTS["remove_cat_prompt"], reply_markup=InlineKeyboardMarkup(keyboard))
            return
        elif data.startswith('cat_remove_final_'):
            parts = data.split('_')
            cat_type = parts[3]
            cat_name = '_'.join(parts[4:])

            if self.data_manager.remove_custom_category(user_id, cat_type, cat_name):
                await query.edit_message_text(TEXTS["cat_remove_success"].format(name=cat_name))
            else:
                await query.edit_message_text(TEXTS["cat_remove_fail"].format(name=cat_name))
            return

        # --- Analytics Callbacks ---
        elif data.startswith('analytics_'):
            if not self.chart_manager.PLOTLY_AVAILABLE:
                await query.edit_message_text("❌ Chart ပြုလုပ်ရန် လိုအပ်သော Library မရှိပါ။ (Plotly)")
                return

            if not self.data_manager.get_premium_status(user_id)['is_premium']:
                await query.answer("🚫 Premium မရှိသေးပါ။", show_alert=True)
                return

            await query.edit_message_text(TEXTS["processing"])

            parts = data.split('_')
            analytics_type = parts[1]
            chart_type = parts[3]

            today = dt.datetime.now()
            month_date = dt.datetime(today.year, today.month, 1)

            transactions = self.data_manager.get_transactions(
                user_id, start_date=month_date)

            if not transactions:
                await query.edit_message_text(TEXTS["no_data"])
                return

            df = pd.DataFrame(transactions)
            df_filtered = df[df['type'] == analytics_type]

            if df_filtered.empty:
                title_my = "ထွက်ငွေ" if analytics_type == 'expense' else "ဝင်ငွေ"
                await query.edit_message_text(f"ℹ️ ယခုလတွင် {title_my} မှတ်တမ်း မရှိသေးပါ။")
                return

            title_my = f"{'ထွက်ငွေ' if analytics_type == 'expense' else 'ဝင်ငွေ'} ခွဲခြမ်းစိတ်ဖြာချက် ({chart_type.upper()} Chart)"
            title = f"{format_myanmar_date(month_date)} - {title_my}"

            chart_buffer = None
            if chart_type == 'pie':
                chart_buffer = self.chart_manager.create_category_pie_chart(
                    df_filtered, title)
            else:
                chart_buffer = self.chart_manager.create_category_bar_chart(
                    df_filtered, title)

            if chart_buffer:
                await query.message.reply_photo(photo=chart_buffer, caption=title)
                await query.edit_message_text(f"✅ {title} ကို အပေါ်တွင် ပုံဖြင့် ဖော်ပြထားပါသည်။", reply_markup=None)
            else:
                await query.edit_message_text("❌ Chart ပြုလုပ်ရာတွင် အမှားဖြစ်ပွားပါသည်။", reply_markup=None)

        # === UPDATED PREMIUM CALLBACK FLOW (STEP 3) ===
        elif data == 'premium_menu_back':
            status = self.data_manager.get_premium_status(user_id)
            if not status['is_premium']:
                message_text = TEXTS['premium_menu_header'] + "\n\n" + \
                    TEXTS['premium_menu_content'] + \
                    "\n\n" + TEXTS['premium_paywall']
                keyboard = [
                    [InlineKeyboardButton(
                        "⭐️ Premium Plan ယူရန်", callback_data='premium_0')],
                    [InlineKeyboardButton(
                        "🎁 ၇ ရက် Free Trial ယူရန်", callback_data='premium_1')]
                ]
                reply_markup = InlineKeyboardMarkup(keyboard)
                await query.edit_message_text(message_text, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup)
            return

        elif data.startswith('premium_'):
            query_data_parts = data.split('_')
            action = query_data_parts[1]

            status = self.data_manager.get_premium_status(user_id)
            if status['is_premium'] and action not in ['menu', 'back']:
                await query.edit_message_text(f"✅ သင့် Premium/Trial မှာ {status['end_date']} နေ့အထိ သုံးစွဲခွင့်ရှိနေပါပြီ။", reply_markup=None)
                return

            if action == '0':
                keyboard = [
                    [InlineKeyboardButton(
                        TEXTS["premium_duration_1"], callback_data='premium_duration_1')],
                    [InlineKeyboardButton(
                        TEXTS["premium_duration_6"], callback_data='premium_duration_6')],
                    [InlineKeyboardButton(
                        TEXTS["premium_duration_12"], callback_data='premium_duration_12')],
                    [InlineKeyboardButton(
                        "↩️ နောက်သို့", callback_data='premium_menu_back')]
                ]
                await query.edit_message_text(
                    TEXTS["premium_select_duration"],
                    parse_mode=ParseMode.MARKDOWN,
                    reply_markup=InlineKeyboardMarkup(keyboard)
                )
                return

            elif action == '1':
                if status['used_trial']:
                    await query.edit_message_text("❌ Free Trial ကို တစ်ကြိမ်သာ သုံးစွဲနိုင်ပါသည်။ ကျေးဇူးပြု၍ ပုံမှန် Premium Plan ကို ယူပေးပါ။", reply_markup=None)
                    return

                end_date = self.data_manager.grant_premium(
                    user_id, 7, is_trial=True)
                await query.edit_message_text(
                    TEXTS["premium_trial_granted"].format(end_date=end_date),
                    parse_mode=ParseMode.MARKDOWN,
                    reply_markup=None
                )
                return

            elif action == 'duration':
                if len(query_data_parts) < 3:
                    return
                months = query_data_parts[2]

                duration_text = TEXTS.get(
                    f"premium_duration_{months}", f"{months} လ")
                price_text = TEXTS.get(f"premium_price_{months}", "N/A")

                keyboard = [
                    [InlineKeyboardButton(
                        TEXTS["premium_payment_wave"], callback_data=f'premium_payment_{months}_wave')],
                    [InlineKeyboardButton(
                        TEXTS["premium_payment_kpay"], callback_data=f'premium_payment_{months}_kpay')],
                    [InlineKeyboardButton(
                        TEXTS["premium_payment_aya"], callback_data=f'premium_payment_{months}_aya')],
                    [InlineKeyboardButton(
                        "↩️ နောက်သို့", callback_data='premium_0')]
                ]

                message_text = TEXTS["premium_select_payment"].format(
                    duration_text=duration_text,
                    price_text=price_text
                )

                await query.edit_message_text(
                    message_text,
                    parse_mode=ParseMode.MARKDOWN,
                    reply_markup=InlineKeyboardMarkup(keyboard)
                )
                return

            elif action == 'payment':
                if len(query_data_parts) < 4:
                    return
                months = query_data_parts[2]
                method = query_data_parts[3]

                price_text = TEXTS.get(f"premium_price_{months}", "N/A")
                payment_title = TEXTS.get(
                    f"premium_payment_{method}", method.upper())
                payment_details = TEXTS.get(
                    f"premium_payment_details_{method}", "Admin အကောင့် အချက်အလက် မရှိပါ။")

                message_text = TEXTS["premium_final_instructions_admin_link_removed"].format(
                    payment_title=payment_title,
                    payment_details=payment_details,
                    payment_method=payment_title,
                    price_text=price_text
                )

                keyboard = [
                    [InlineKeyboardButton(
                        TEXTS["premium_payment_complete_button"], callback_data=f'premium_paid_{months}')],
                    [InlineKeyboardButton(
                        "↩️ Payment ပြန်ရွေးရန်", callback_data=f'premium_duration_{months}')]
                ]

                await query.edit_message_text(
                    message_text,
                    parse_mode=ParseMode.MARKDOWN,
                    disable_web_page_preview=True,
                    reply_markup=InlineKeyboardMarkup(keyboard)
                )
                return

            elif action == 'paid':
                if len(query_data_parts) < 3:
                    return
                months = query_data_parts[2]

                context.user_data['mode'] = 'awaiting_screenshot'
                context.user_data['plan_months'] = months

                await query.edit_message_text(
                    TEXTS["premium_awaiting_screenshot"],
                    parse_mode=ParseMode.MARKDOWN,
                    reply_markup=None
                )
                return

        # === END OF UPDATED PREMIUM FLOW ===

        elif data.startswith('select_month_'):
            if not self.data_manager.get_premium_status(user_id)['is_premium']:
                await query.answer("🚫 Premium မရှိသေးပါ။", show_alert=True)
                return

            month_str = data.split('_')[2]
            context.user_data.update(
                {'mode': 'monthly_report', 'report_month': month_str})

            export_options = [[InlineKeyboardButton("PDF (.pdf)", callback_data='export_type_monthly_pdf')], [
                InlineKeyboardButton("Excel (.xlsx)", callback_data='export_type_monthly_xlsx')]]
            await context.bot.send_message(user_id, TEXTS["export_select_type"], reply_markup=InlineKeyboardMarkup(export_options))
            try:
                selected_month_text = next(
                    (button.text for row in query.message.reply_markup.inline_keyboard for button in row if button.callback_data == data), month_str)
                await query.edit_message_text(f"✅ လရွေးချယ်မှု: {selected_month_text} အတွက် အစီရင်ခံစာ အမျိုးအစားကို ရွေးချယ်ပါ။", reply_markup=None)
            except Exception:
                pass

        elif data.startswith('export_type_monthly_'):
            if not self.data_manager.get_premium_status(user_id)['is_premium']:
                await query.answer("🚫 Premium မရှိသေးပါ။", show_alert=True)
                return

            state = context.user_data

            if state.get('mode') != 'monthly_report' or 'report_month' not in state:
                await context.bot.send_message(user_id, "❌ လရွေးချယ်မှု ပျက်သွားပါပြီ။ 🧾 အစီရင်ခံစာ ကို ပြန်နှိပ်ပါ။")
                return

            export_type = data.split('_')[3]
            month_str = state['report_month']

            await query.edit_message_text(f'⏳ {month_str} အတွက် {export_type.upper()} အစီရင်ခံစာကို ပြုလုပ်နေပါသည်။ ခဏစောင့်ပါ။')

            try:
                year, month = map(int, month_str.split('-'))
                month_date = dt.datetime(year, month, 1)
            except ValueError:
                await context.bot.send_message(user_id, "❌ မမှန်ကန်သော ရက်စွဲဖြစ်ပါသည်။")
                return

            await self.send_report(user_id, month_date, export_type, context)

            context.user_data.clear()

        elif data.startswith('export_custom_'):
            if not self.data_manager.get_premium_status(user_id)['is_premium']:
                await query.answer("🚫 Premium မရှိသေးပါ။", show_alert=True)
                return

            state = context.user_data

            if state.get('mode') != 'custom_report' or 'start_date' not in state or 'end_date' not in state:
                await context.bot.send_message(user_id, "❌ ရက်စွဲရွေးချယ်မှု ပျက်သွားပါပြီ။ Premium Menu မှ Custom Report ကို ပြန်နှိပ်ပါ။")
                return

            export_type = data.split('_')[2]
            start_date_iso = state['start_date']
            end_date_iso = state['end_date']

            await query.edit_message_text(f"⏳ {dt.datetime.fromisoformat(start_date_iso).strftime('%Y-%m-%d')} မှ {dt.datetime.fromisoformat(end_date_iso).strftime('%Y-%m-%d')} အထိ {export_type.upper()} အစီရင်ခံစာကို ပြုလုပ်နေပါသည်။ ခဏစောင့်ပါ။")

            start_date = dt.datetime.fromisoformat(start_date_iso)
            end_date = dt.datetime.fromisoformat(end_date_iso)

            await self.send_report(user_id, start_date, export_type, context, end_date=end_date)

            context.user_data.clear()

        # --- Reminder Callbacks ---
        elif data == 'weekly_reminder_select_day':
            if not self.data_manager.get_premium_status(user_id)['is_premium']:
                return

            settings = self.data_manager.get_reminder_settings(user_id)
            current_day = settings.get('weekly_day', 'Sunday')

            keyboard = []
            for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']:
                status_emoji = '☑️' if day == current_day else '⬜️'
                keyboard.append([InlineKeyboardButton(
                    f"{status_emoji} {day}", callback_data=f'set_reminder_day_{day}')])

            weekly_status_text = "❌ ပိတ်ရန်" if settings.get(
                'weekly_summary') else "✅ ဖွင့်ရန်"
            keyboard.append([InlineKeyboardButton(
                weekly_status_text, callback_data='toggle_weekly_reminder')])
            keyboard.append([InlineKeyboardButton(
                "↩️ သတိပေးချက် မီနူးသို့", callback_data='reminder_menu_back')])

            await query.edit_message_text("အပတ်စဉ် အစီရင်ခံစာ ပို့ပေးမည့်နေ့ကို ရွေးချယ်ပါ၊ သို့မဟုတ် ဖွင့်/ပိတ် လုပ်ပါ။", reply_markup=InlineKeyboardMarkup(keyboard))
            return

        elif data.startswith('set_reminder_day_'):
            if not self.data_manager.get_premium_status(user_id)['is_premium']:
                return
            day = data.split('_')[3]

            self.data_manager.set_reminder_setting(
                user_id, 'weekly_summary', True)
            self.data_manager.set_reminder_setting(user_id, 'weekly_day', day)

            await query.edit_message_text(TEXTS["reminder_set_success"].format(day=day))
            await self.reminder_menu(update, context)
            return

        elif data == 'toggle_weekly_reminder':
            if not self.data_manager.get_premium_status(user_id)['is_premium']:
                return

            current_status = self.data_manager.get_reminder_settings(
                user_id).get('weekly_summary', False)
            new_status = not current_status

            self.data_manager.set_reminder_setting(
                user_id, 'weekly_summary', new_status)

            message = f"✅ အပတ်စဉ် အစီရင်ခံစာ သတိပေးချက်ကို ဖွင့်လိုက်ပါပြီ။" if new_status else f"❌ အပတ်စဉ် အစီရင်ခံစာ သတိပေးချက်ကို ပိတ်လိုက်ပါပြီ။"
            await query.edit_message_text(message)
            await self.reminder_menu(update, context)
            return

        elif data == 'toggle_daily_reminder':
            if not self.data_manager.get_premium_status(user_id)['is_premium']:
                return

            current_status = self.data_manager.get_reminder_settings(
                user_id).get('daily_transaction', False)
            new_status = not current_status

            self.data_manager.set_reminder_setting(
                user_id, 'daily_transaction', new_status)

            message = TEXTS["daily_reminder_on"] if new_status else TEXTS["daily_reminder_off"]
            await query.edit_message_text(message)
            await self.reminder_menu(update, context)
            return

        elif data == 'reminder_menu_back':
            await self.reminder_menu(update, context)
            return

    # --- send_report method ---

    async def send_report(self, user_id: int, start_date: dt.datetime, export_type: str, context: ContextTypes.DEFAULT_TYPE, end_date: Optional[dt.datetime] = None):

        chart_src = None

        if end_date is None:
            month_date = start_date
            month_str_my = format_myanmar_date(month_date)

            if month_date.month == 12:
                report_end_date = dt.datetime(
                    month_date.year + 1, 1, 1) - timedelta(days=1)
            else:
                report_end_date = dt.datetime(
                    month_date.year, month_date.month + 1, 1) - timedelta(days=1)

            report_start_date = dt.datetime(
                month_date.year, month_date.month, 1)
            title = f"{month_str_my} လစဉ် ငွေကြေးအစီရင်ခံစာ"
            caption_text = TEXTS["export_success"].format(
                month=month_str_my, type=export_type.upper())
            file_name = f"monthly_report_{month_date.strftime('%Y_%m')}.{export_type}"

        else:
            report_start_date = start_date
            report_end_date = end_date

            start_str = report_start_date.strftime('%Y-%m-%d')
            end_str = report_end_date.strftime('%Y-%m-%d')
            title = f"ငွေကြေးအစီရင်ခံစာ ({start_str} မှ {end_str} အထိ)"
            caption_text = f"✅ {start_str} မှ {end_str} အထိ {export_type.upper()} အစီရင်ခံစာကို အောက်ပါအတိုင်း ထုတ်ယူပေးလိုက်ပါပြီ။"
            file_name = f"custom_report_{report_start_date.strftime('%Y%m%d')}_{report_end_date.strftime('%Y%m%d')}.{export_type}"

        transactions = self.data_manager.get_transactions(
            user_id, start_date=report_start_date, end_date=report_end_date)

        if not transactions:
            await context.bot.send_message(user_id, TEXTS["data_not_found"])
            return

        for tx in transactions:
            if isinstance(tx['date'], str):
                tx['date'] = self._parse_date(tx['date'])

        if export_type == 'pdf' and self.chart_manager.PLOTLY_AVAILABLE:
            df = pd.DataFrame(transactions)
            df_expense = df[df['type'] == 'expense']

            if not df_expense.empty:
                chart_title = f"{title} - အသုံးစရိတ် Chart"
                chart_buffer = self.chart_manager.create_category_pie_chart(
                    df_expense, chart_title)

                if chart_buffer:
                    try:
                        chart_buffer.seek(0)
                        chart_base64 = base64.b64encode(
                            chart_buffer.read()).decode('utf-8')
                        chart_src = f"data:image/png;base64,{chart_base64}"
                        logger.info("Chart generated and encoded for PDF.")
                    except Exception as e:
                        logger.error(f"Failed to encode chart for PDF: {e}")

        file_buffer = self.export_manager.export_data(
            title, transactions, export_type, chart_data=chart_src)

        if file_buffer:
            await context.bot.send_document(
                chat_id=user_id,
                document=file_buffer,
                filename=file_name,
                caption=caption_text,
                read_timeout=30.0,
                write_timeout=30.0,
                pool_timeout=30.0
            )
        else:
            await context.bot.send_message(user_id, TEXTS["export_failure"])

    # --- Send Daily Reminder ---
    async def send_daily_transaction_reminder(self, user_id: int, time_of_day: str):
        try:
            context = ContextTypes.DEFAULT_TYPE(application=self.application)

            if time_of_day == 'morning':
                message = TEXTS['daily_reminder_morning']
            else:
                message = TEXTS['daily_reminder_evening']

            keyboard = [[KeyboardButton(text) for text in row]
                        for row in TEXTS["main_reply_buttons"][:1]]
            reply_markup = ReplyKeyboardMarkup(
                keyboard, resize_keyboard=True, one_time_keyboard=False)

            await context.bot.send_message(
                user_id,
                message,
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=reply_markup
            )
        except Exception as e:
            logger.error(f"Failed to send daily reminder to {user_id}: {e}")

    # --- Scheduler Methods ---
    def setup_reminders(self):
        self.scheduler.add_job(self._check_and_send_reminders,
                               'cron', hour=9, minute=0, name='Weekly_Summary_Check')
        self.scheduler.add_job(self._check_and_send_reminders,
                               'cron', hour=9, minute=0, name='Daily_Tx_Morning_Check')
        self.scheduler.add_job(self._check_and_send_reminders,
                               'cron', hour=19, minute=0, name='Daily_Tx_Evening_Check')
        self.scheduler.add_job(self._check_and_run_recurring_tx,
                               'cron', hour=8, minute=0, name='Recurring_TX_Check')

    async def _check_and_send_reminders(self):
        today = dt.datetime.now()
        day_name = today.strftime('%A')
        current_hour = today.hour

        users_to_remind = self.data_manager.get_all_users_for_reminders()

        for user_id, daily_on, weekly_day, weekly_on in users_to_remind:
            try:
                if weekly_on and weekly_day.capitalize() == day_name and current_hour == 9:
                    logger.info(
                        f"Sending weekly summary reminder to user {user_id} on {day_name}")
                    await self.send_weekly_summary(user_id)

                if daily_on:
                    if current_hour == 9:
                        logger.info(
                            f"Sending daily transaction morning reminder to user {user_id}")
                        await self.send_daily_transaction_reminder(user_id, 'morning')
                    elif current_hour == 19:
                        logger.info(
                            f"Sending daily transaction evening reminder to user {user_id}")
                        await self.send_daily_transaction_reminder(user_id, 'evening')

            except Exception as e:
                logger.error(
                    f"Error checking reminder for user {user_id}: {e}")

    # --- Scheduler job for Recurring Transactions ---
    async def _check_and_run_recurring_tx(self):
        today_day = dt.datetime.now().day
        logger.info(
            f"Running recurring transaction check for day {today_day}...")

        context = ContextTypes.DEFAULT_TYPE(application=self.application)

        user_ids = self.data_manager.get_all_users_for_recurring_tx()

        for user_id in user_ids:
            try:
                recurring_txs = self.data_manager.get_recurring_txs(user_id)

                for rtx in recurring_txs:
                    if rtx['day'] == today_day:
                        today_start = dt.datetime.now().replace(
                            hour=0, minute=0, second=0, microsecond=0)
                        txs_today = self.data_manager.get_transactions(
                            user_id, start_date=today_start, end_date=dt.datetime.now())

                        already_added = False
                        for tx in txs_today:
                            if tx['description'] == rtx['description'] and tx['amount'] == rtx['amount'] and tx['type'] == rtx['type']:
                                already_added = True
                                break

                        if not already_added:
                            logger.info(
                                f"Executing recurring transaction '{rtx['description']}' for user {user_id}")
                            self.data_manager.add_transaction(
                                user_id,
                                rtx['type'],
                                rtx['amount'],
                                rtx['description'],
                                rtx['category']
                            )
                            await context.bot.send_message(
                                user_id,
                                TEXTS["recurring_tx_executed"].format(
                                    desc=rtx['description'],
                                    amount=rtx['amount']
                                ),
                                parse_mode=ParseMode.MARKDOWN
                            )
                        else:
                            logger.warning(
                                f"Skipping recurring tx '{rtx['description']}' for user {user_id}, already added today.")

            except Exception as e:
                logger.error(
                    f"Error processing recurring transaction for user {user_id}: {e}")

    async def send_weekly_summary(self, user_id: int):
        class MockUpdate:
            def __init__(self, uid):
                self.effective_user = type(
                    'User', (), {'id': uid, 'mention_html': lambda: f"User {uid}"})()
                self.message = type('Message', (), {'text': '/summary'})()

        try:
            update = MockUpdate(user_id)
            context = ContextTypes.DEFAULT_TYPE(application=self.application)
            await context.bot.send_message(user_id, "🔔 **အပတ်စဉ် အစီရင်ခံစာ သတိပေးချက်**\n\nယခုတစ်ပတ်၏ ငွေစာရင်းအခြေအနေကို အောက်ပါအတိုင်း စစ်ဆေးနိုင်ပါပြီ။", parse_mode=ParseMode.MARKDOWN)
            await self.summary(update, context)
        except Exception as e:
            logger.error(f"Failed to send weekly summary to {user_id}: {e}")

     # --- (STEP 5.1) NEW: AI Financial Analyst Handler ---
    async def run_ai_analysis(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id

        try:
            analysis_data = self.data_manager.get_financial_analysis_data(
                user_id)

            if not analysis_data or (analysis_data['total_income'] == 0 and analysis_data['total_expense'] == 0):
                await context.bot.send_message(user_id, TEXTS["ai_analysis_no_data"], parse_mode=ParseMode.MARKDOWN)
                return

            insights = []  # AI က တွေ့ရှိချက်တွေကို ဒီ list ထဲ ထည့်ပါမယ်

            income = analysis_data['total_income']
            expense = analysis_data['total_expense']
            expense_breakdown = analysis_data['expense_breakdown']
            budgets = analysis_data['budgets']

            # --- Rule 1: Saving Rate (စုငွေနှုန်း) ---
            if income > 0:
                saving_rate = ((income - expense) / income) * 100
                if saving_rate < 0:
                    insights.append(TEXTS["ai_insight_saving_rate_negative"].format(
                        expense=expense, income=income))
                elif saving_rate < 15:
                    insights.append(
                        TEXTS["ai_insight_saving_rate_low"].format(rate=saving_rate))
                else:
                    insights.append(
                        TEXTS["ai_insight_saving_rate_good"].format(rate=saving_rate))
            elif expense > 0:
                # ဝင်ငွေ မရှိဘဲ ထွက်ငွေပဲ ရှိရင်
                insights.append(TEXTS["ai_insight_saving_rate_negative"].format(
                    expense=expense, income=0))

            # --- Rule 2: Top Expense Category (အသုံးအများဆုံး Category) ---
            if expense > 0 and expense_breakdown:
                # Category တွေကို သုံးစွဲမှုအလိုက် စီပါ
                sorted_expenses = sorted(
                    expense_breakdown.items(), key=lambda item: item[1], reverse=True)

                # အများဆုံး သုံးတဲ့ တစ်ခု
                top_cat, top_cat_amount = sorted_expenses[0]
                top_cat_percent = (top_cat_amount / expense) * 100

                # စုစုပေါင်း ထွက်ငွေရဲ့ 30% ထက် ပိုသုံးထားမှသာ သတိပေးပါ
                if top_cat_percent > 30:
                    insights.append(TEXTS["ai_insight_top_expense"].format(
                        percent=top_cat_percent, category=top_cat))

            # --- Rule 3: Budget Check (ဘတ်ဂျက် စစ်ဆေးခြင်း) ---
            if budgets and expense_breakdown:
                for category, budgeted_amount in budgets.items():
                    spent_amount = expense_breakdown.get(category, 0)

                    if spent_amount > 0 and budgeted_amount > 0:
                        percent_spent = (spent_amount / budgeted_amount) * 100

                        if percent_spent >= 100:
                            # ဘတ်ဂျက် ကျော်လွန်သွားပြီ
                            insights.append(TEXTS["ai_insight_budget_over"].format(
                                category=category,
                                budget=budgeted_amount,
                                percent=percent_spent,
                                spent=spent_amount
                            ))
                        elif percent_spent >= 80:
                            # ဘတ်ဂျက် 80% နီးပါး ရောက်ပြီ
                            insights.append(TEXTS["ai_insight_budget_warning"].format(
                                category=category,
                                budget=budgeted_amount,
                                percent=percent_spent,
                                spent=spent_amount
                            ))

            # --- Final Message ---
            if not insights:
                # အပေါ်က Rule တွေ တစ်ခုမှ မမိရင် (ဥပမာ- ဝင်ငွေပဲ ရှိပြီး ထွက်ငွေ မရှိရင်)
                final_message = "✅ သင်၏ ငွေစာရင်းမှာ ပုံမှန် အခြေအနေတွင် ရှိပါသည်။"
            else:
                final_message = TEXTS["ai_analysis_header"] + "".join(insights)

            # Message အသစ် ပို့ပေးပါ (Edit မလုပ်ပါ)
            await context.bot.send_message(user_id, final_message, parse_mode=ParseMode.MARKDOWN)

        except Exception as e:
            logger.error(f"Error during AI analysis for user {user_id}: {e}")
            await context.bot.send_message(user_id, "❌ AI သုံးသပ်ချက် ပြုလုပ်ရာတွင် အမှားဖြစ်ပွားပါသည်။")

    # --- Manage Transactions Menu ---
    async def manage_transactions_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        recent_txs = self.data_manager.get_recent_transactions(
            user_id, limit=5)
        is_premium = self.data_manager.get_premium_status(user_id)[
            'is_premium']

        keyboard = []
        
        # --- (!!!) NEW: Account Management Button (!!!) ---
        # Multi-Wallet feature ကို အပေါ်ဆုံးမှာ ထားပါ
        keyboard.append([
            InlineKeyboardButton("💰 Account စီမံခန့်ခွဲ", callback_data='account_menu')
        ])
        # --- (!!!) End of New (!!!) ---

        if is_premium:
            keyboard.append([
                InlineKeyboardButton(TEXTS["backup_restore_button"], callback_data='backup_restore_menu'),
                InlineKeyboardButton(TEXTS["info_button_text"], callback_data='info_backup_restore')
            ])
        # ... (ကျန်တဲ့ code တွေ ဒီအတိုင်း ဆက်ထားပါ)
            keyboard.append([
                InlineKeyboardButton("🔁 လစဉ် ထပ်တလဲလဲ (Recurring)", callback_data='recurring_tx_menu'),
                InlineKeyboardButton(TEXTS["info_button_text"], callback_data='info_recurring_tx') # <-- ထည့်ရန်
            ])

        if not recent_txs:
            if not is_premium:
                await context.bot.send_message(user_id, TEXTS["no_recent_tx"])
                return
        else:
            for tx in recent_txs:
                tx_date_obj = self._parse_date(tx['date'])
                if tx_date_obj:
                    tx_date = tx_date_obj.strftime('%m/%d')
                else:
                    tx_date = "N/A"

                tx_type_my = "ဝင်ငွေ" if tx['type'] == 'income' else "ထွက်ငွေ"
                tx_label = f"{tx_date} - {tx_type_my} ({tx['category']}) : {tx['amount']:,.0f} Ks"
                keyboard.append([InlineKeyboardButton(
                    tx_label, callback_data=f'tx_select_{tx["id"]}')])

        message_text = TEXTS['manage_tx_menu']
        if recent_txs:
            message_text += "\n\n" + TEXTS['select_tx_action']

        if update.callback_query:
            try:
                await update.callback_query.edit_message_text(
                    message_text,
                    parse_mode=ParseMode.MARKDOWN,
                    reply_markup=InlineKeyboardMarkup(keyboard)
                )
            except Exception as e:
                logger.warning(
                    f"Failed to edit message for manage_transactions_menu: {e}")
                await context.bot.send_message(
                    user_id,
                    message_text,
                    parse_mode=ParseMode.MARKDOWN,
                    reply_markup=InlineKeyboardMarkup(keyboard)
                )
        else:
            await context.bot.send_message(
                user_id,
                message_text,
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
    # (!!!) --- NEW: Account Management Bot Functions --- (!!!)
    
    async def account_menu(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Displays the main account management menu."""
        keyboard = [
            [InlineKeyboardButton(TEXTS["account_view_button"], callback_data='account_view')],
            [InlineKeyboardButton(TEXTS["account_add_button"], callback_data='account_add_start')],
            # [InlineKeyboardButton(TEXTS["account_transfer_button"], callback_data='account_transfer_start')] # <-- ဒါကို နောက် အဆင့်မှ ထည့်ပါမယ်
            [InlineKeyboardButton("↩️ စီမံခန့်ခွဲ မီနူးသို့", callback_data='manage_tx_menu_back')]
        ]
        
        message_text = TEXTS["account_menu_header"]
        
        if update.callback_query:
            await update.callback_query.edit_message_text(message_text, parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(keyboard))
        else:
            await context.bot.send_message(update.effective_user.id, message_text, parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(keyboard))

    async def account_view_balances(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Shows a list of all accounts and their balances."""
        user_id = update.effective_user.id
        
        if update.callback_query:
            await update.callback_query.answer("💰 လက်ကျန်ငွေများ စစ်ဆေးနေပါသည်...")

        accounts_with_balance = self.data_manager.get_accounts_with_balance(user_id)
        
        if not accounts_with_balance:
            message_text = TEXTS["account_list_empty"]
        else:
            message_text = TEXTS["account_list_header"]
            total_balance = 0
            
            for acc in accounts_with_balance:
                message_text += TEXTS["account_list_detail"].format(name=acc['name'], balance=acc['balance'])
                total_balance += acc['balance']
            
            message_text += TEXTS["account_list_total"].format(total=total_balance)

        keyboard = [
            [InlineKeyboardButton(TEXTS["account_add_button"], callback_data='account_add_start')],
            [InlineKeyboardButton("↩️ Account မီနူးသို့", callback_data='account_menu')]
        ]

        if update.callback_query:
            await update.callback_query.edit_message_text(message_text, parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(keyboard))
        else:
            await context.bot.send_message(user_id, message_text, parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(keyboard))

    async def account_add_prompt(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Sets the state to wait for new account name and balance."""
        context.user_data['mode'] = 'add_account'
        await update.callback_query.edit_message_text(TEXTS["account_add_prompt"], parse_mode=ParseMode.MARKDOWN)

    # (!!!) --- End of New Account Functions --- (!!!)
    # (!!!) --- NEW: Budget Alert Helper Function --- (!!!)
    async def check_budget_alert(self, user_id: int, category: str, amount: int, context: ContextTypes.DEFAULT_TYPE):
        """Checks and sends a budget alert if needed."""
        try:
            budgets = self.data_manager.get_budgets(user_id)
            if category in budgets:
                current_tx = {'type': 'expense', 'amount': amount, 'category': category}
                _, _, alert_needed = self.calculate_budget_status(
                    user_id, current_tx=current_tx)

                if alert_needed:
                    budgeted_amount = budgets[category]
                    today = dt.datetime.now()
                    start_of_month = today.replace(
                        day=1, hour=0, minute=0, second=0)
                    transactions = self.data_manager.get_transactions(
                        user_id, start_date=start_of_month)
                    expense_df = pd.DataFrame(
                        [tx for tx in transactions if tx['type'] == 'expense'])
                    spent = expense_df[expense_df['category']
                                       == category]['amount'].sum()
                    remaining = budgeted_amount - spent
                    percent_spent = (spent / budgeted_amount) * 100

                    alert_message = TEXTS["budget_alert_overrun"].format(
                        category=category,
                        budget=budgeted_amount,
                        percent=percent_spent,
                        spent=spent,
                        remaining=remaining
                    )
                    await context.bot.send_message(user_id, alert_message, parse_mode=ParseMode.MARKDOWN)
        except Exception as e:
            logger.error(f"Failed to check budget alert for user {user_id}: {e}")
    # (!!!) --- End of New Helper --- (!!!)
    # (!!!) --- NEW: Helper Function for Account Selection --- (!!!)
    async def prompt_account_selection(self, update_or_message: Update | Any, context: ContextTypes.DEFAULT_TYPE, user_id: int, prompt_text: str):
        """
        User ကို Account ရွေးခိုင်းတဲ့ ခလုတ်တွေ (Keyboard) ကို ပြပေးမယ့် Helper Function
        """
        accounts = self.data_manager.get_accounts(user_id) # <-- (!!!) ဒါက အခု List[Dict] ဖြစ်သွားပါပြီ
        
        # Account မရှိသေးရင်၊ Account အရင် ဆောက်ခိုင်းပါ
        if not accounts:
            await context.bot.send_message(user_id, TEXTS["no_accounts_error"], parse_mode=ParseMode.MARKDOWN)
            context.user_data.clear() # State ကို ရှင်းလင်းပါ
            return

        keyboard = []
        row = []
        for acc in accounts:
            # (!!!) Object (acc.name) အစား၊ Dict Key (acc['name']) ကို သုံးပါ (!!!)
            acc_name = acc['name']
            acc_id = acc['id']
            row.append(InlineKeyboardButton(f"💰 {acc_name}", callback_data=f'tx_select_account_{acc_id}'))
            if len(row) == 2: # တစ်တန်းမှာ ၂ ခု
                keyboard.append(row)
                row = []
        if row:
            keyboard.append(row)
            
        # "Account မသတ်မှတ်" ဆိုတဲ့ ခလုတ်ကို အောက်ဆုံးမှာ ထည့်ပေးပါ
        keyboard.append([InlineKeyboardButton(TEXTS["select_account_button_none"], callback_data='tx_select_account_none')])

        reply_markup = InlineKeyboardMarkup(keyboard)

        # Message ကို Edit လုပ်မလား (Quick Add) / Message အသစ် ပို့မလား (Normal Add)
        if isinstance(update_or_message, Update) and update_or_message.callback_query:
            # This is a callback query (from Quick Add), so edit the message
            await update_or_message.callback_query.edit_message_text(prompt_text, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup)
        else:
            # This is a new message (from Normal Add), so send a new message
            await context.bot.send_message(user_id, prompt_text, parse_mode=ParseMode.MARKDOWN, reply_markup=reply_markup)
    # (!!!) --- End of New Helper Function --- (!!!)
# --- (STEP 4) NEW: Admin Dashboard Handlers ---

    async def admin_dashboard(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        user_id = update.effective_user.id
        if user_id != self.ADMIN_ID:
            if update.callback_query:
                return await update.callback_query.answer(TEXTS["not_admin"], show_alert=True)
            elif update.message:
                return await update.message.reply_text(TEXTS["not_admin"])
            return

        keyboard = [
            [InlineKeyboardButton(
                TEXTS["admin_stats_button"], callback_data='admin_stats')],
            [InlineKeyboardButton(
                TEXTS["admin_broadcast_button"], callback_data='admin_broadcast_prompt')],
            [InlineKeyboardButton(
                TEXTS["admin_find_user_button"], callback_data='admin_find_user_prompt')]
        ]

        message_text = TEXTS["admin_dashboard_header"]
        reply_markup = InlineKeyboardMarkup(keyboard)

        if update.callback_query:
            # ခလုတ်ကနေ ခေါ်တာ (ဥပမာ: "Back" button) ဆိုရင် message ကို edit လုပ်ပါ
            await update.callback_query.edit_message_text(
                message_text,
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=reply_markup
            )
        elif update.message:
            # /admin command (text) ကနေ ခေါ်တာဆိုရင် message အသစ် ပို့ပါ
            await update.message.reply_text(
                message_text,
                parse_mode=ParseMode.MARKDOWN,
                reply_markup=reply_markup
            )

    async def admin_stats(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        stats = self.data_manager.get_stats()
        message = TEXTS["admin_stats_message"].format(
            total=stats.get('total', 0),
            premium=stats.get('premium', 0)
        )
        keyboard = [[InlineKeyboardButton(
            "↩️ Admin Menu သို့ ပြန်သွားရန်", callback_data='admin_dashboard')]]

        await update.callback_query.edit_message_text(
            message,
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    async def admin_broadcast_prompt(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        context.user_data['mode'] = 'admin_broadcast_message'
        await update.callback_query.edit_message_text(
            TEXTS["admin_broadcast_prompt"],
            parse_mode=ParseMode.MARKDOWN
        )

    async def _run_broadcast(self, admin_chat_id: int, message: str, context: ContextTypes.DEFAULT_TYPE):
        """Helper async task to send broadcast messages without blocking."""
        user_ids = self.data_manager.get_all_user_ids()
        count = len(user_ids)
        sent_count = 0
        failed_count = 0

        await context.bot.send_message(
            admin_chat_id,
            TEXTS["admin_broadcast_start"].format(count=count),
            parse_mode=ParseMode.MARKDOWN
        )

        for user_id in user_ids:
            if user_id == admin_chat_id:  # Don't send to self
                sent_count += 1
                continue
            try:
                await context.bot.send_message(user_id, message, parse_mode=ParseMode.MARKDOWN)
                sent_count += 1
            except Exception as e:
                logger.warning(f"Broadcast failed for user {user_id}: {e}")
                failed_count += 1

            await asyncio.sleep(0.1)  # Avoid rate limits

        await context.bot.send_message(
            admin_chat_id,
            TEXTS["admin_broadcast_complete"].format(
                sent=sent_count, failed=failed_count),
            parse_mode=ParseMode.MARKDOWN
        )

    async def admin_broadcast_send(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        message = context.user_data.get('broadcast_message')
        if not message:
            await update.callback_query.edit_message_text("❌ Message မရှိပါ။")
            return

        await update.callback_query.edit_message_text("✅ Broadcast ကို စတင် ပို့နေပါပြီ...")

        # Run broadcast in background
        asyncio.create_task(self._run_broadcast(
            update.effective_user.id, message, context))

        context.user_data.clear()

    async def admin_find_user_prompt(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        context.user_data['mode'] = 'admin_find_user_id'
        await update.callback_query.edit_message_text(
            TEXTS["admin_find_user_prompt"],
            parse_mode=ParseMode.MARKDOWN
        )

    async def _send_user_details(self, message: Any, context: ContextTypes.DEFAULT_TYPE, target_user_id: int):
        """Helper to send or edit the user details message."""
        details = self.data_manager.get_user_details(target_user_id)
        if not details:
            # Check if message is callback_query or regular message
            if hasattr(message, 'edit_message_text'):
                return await message.edit_message_text(TEXTS["admin_user_not_found"].format(user_id=target_user_id))
            else:
                return await message.reply_text(TEXTS["admin_user_not_found"].format(user_id=target_user_id))

        status_text = "Active ✅" if details['is_premium'] else "Inactive ❌"

        message_text = TEXTS["admin_user_details"].format(
            id=details['id'],
            status=status_text,
            end_date=details['end_date'],
            tx_count=details['tx_count'],
            used_trial="Yes" if details['used_trial'] else "No"
        )

        keyboard = [
            [InlineKeyboardButton(
                TEXTS["admin_grant_button"], callback_data=f'admin_grant_user_{target_user_id}')],
            [InlineKeyboardButton(TEXTS["admin_revoke_button"],
                                  callback_data=f'admin_revoke_user_{target_user_id}')],
            [InlineKeyboardButton(
                "↩️ Admin Menu သို့ ပြန်သွားရန်", callback_data='admin_dashboard')]
        ]

        if hasattr(message, 'edit_message_text'):  # If called from callback
            await message.edit_message_text(message_text, parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(keyboard))
        else:  # If called from handle_message
            await message.reply_text(message_text, parse_mode=ParseMode.MARKDOWN, reply_markup=InlineKeyboardMarkup(keyboard))
            # --- Post Init Tasks ---

    async def post_init_tasks(self, application: Application):
        """Tasks to run after the bot is initialized but before polling starts."""
        if self.scheduler.state != 1:
            self.scheduler.start()
            logger.info("Scheduler started successfully in PTB event loop.")

    async def run(self):
        """Starts the bot."""
        if not TELEGRAM_BOT_TOKEN:
            print(
                '❌ TELEGRAM_BOT_TOKEN is missing. Please set it in your code or environment variables.')
            return

        if not SQLALCHEMY_AVAILABLE:
            print("❌ SQLAlchemy library is not installed. Bot cannot start.")
            print("Please run: pip install sqlalchemy")
            return

        # --- Persistence (State တွေ မှတ်ထားရန်) ---
        persistence = PicklePersistence(filepath=PERSISTENCE_FILE_PATH)

        self.application = Application.builder().token(
            TELEGRAM_BOT_TOKEN).persistence(persistence).build()

        # --- Handlers (User ဆီက Message တွေကို ဘယ်သူက တာဝန်ယူမလဲ) ---

        # Command Handlers (/)
        self.application.add_handler(CommandHandler('start', self.start))
        self.application.add_handler(CommandHandler('help', self.help))
        self.application.add_handler(CommandHandler('privacy', self.privacy))
        self.application.add_handler(CommandHandler(
            'delete_my_data', self.delete_my_data_command))
        self.application.add_handler(CommandHandler('summary', self.summary))
        self.application.add_handler(CommandHandler(
            'budget_status', self.budget_status))
        self.application.add_handler(CommandHandler(
            'monthly_report', self.monthly_report))
        self.application.add_handler(
            CommandHandler('add_income', self.add_income))
        self.application.add_handler(
            CommandHandler('add_expense', self.add_expense))
        self.application.add_handler(CommandHandler(
            'grant_premium', self.grant_premium_command))

        # Admin Command Handler
        self.application.add_handler(
            CommandHandler('admin', self.admin_dashboard))

        # Message Handlers (ပုံ၊ File၊ စာသား)
        self.application.add_handler(MessageHandler(
            filters.PHOTO & ~filters.COMMAND, self.handle_screenshot))
        self.application.add_handler(MessageHandler(
            filters.Document.ALL, self.handle_backup_file))
        self.application.add_handler(MessageHandler(
            filters.TEXT & ~filters.COMMAND, self.handle_message))

        # Callback Handler (ခလုတ်နှိပ်ခြင်းများ)
        self.application.add_handler(
            CallbackQueryHandler(self.handle_callback))

        # --- Scheduler (အချိန် နဲ့ အလုပ်လုပ်ရန်) ---
        self.setup_reminders()
        self.application.post_init = self.post_init_tasks

        # --- Bot ကို စတင်ခြင်း ---
        print(f'🤖 Myanmar Finance Bot (All Features) is starting...')
        print(f'✅ Bot State Persistence: ENABLED (using bot_persistence file)')
        print(f'✅ Database: ENABLED (using financebot.db)')
        print(f'✅ Admin Approval System: ENABLED')
        print(f'✅ Admin Dashboard: ENABLED (use /admin command)')
        print(f'✅ Quick Add Feature: ENABLED')
        print(f'✅ AI Analyst Feature: ENABLED')
        print(f'✅ Backup/Restore Feature: ENABLED')

        if not WEASYPRINT_AVAILABLE:
            print("⚠️ WeasyPrint is not properly installed. PDF export (and potentially Chart export) will not work.")
        if not PLOTLY_AVAILABLE:
            print("⚠️ Plotly is not installed. Chart generation will be disabled.")

        try:
            # 1. Application ကို initialize လုပ်ပါ (ဒါက post_init ကိုပါ run ပေးပါလိမ့်မယ်)
            await self.application.initialize()
            
            # 2. Polling ကို စတင်ပါ
            await self.application.updater.start_polling(allowed_updates=Update.ALL_TYPES)
            
            # 3. Handler တွေကို စတင် အလုပ်လုပ်ခိုင်းပါ
            await self.application.start()
            
            logger.info("Bot is now running. Press Ctrl-C to stop.")
            
            # Script ဆက် run နေအောင် အကြာကြီး sleep ထားပါ
            while True:
                await asyncio.sleep(3600)
                
        except (KeyboardInterrupt, SystemExit):
            logger.info("Bot stopping (received interrupt)...")
        except Exception as e:
            logger.critical(f"Bot failed to run: {e}", exc_info=True)
        finally:
            # Error တက်ရင် (ဒါမှမဟုတ် Ctrl-C နှိပ်ရင်) အားလုံးကို သေချာ ပြန်ပိတ်ပါ
            if self.application:
                await self.application.stop()
                if self.application.updater:
                    await self.application.updater.stop()
                if self.scheduler.running:
                    self.scheduler.shutdown()
            logger.info("Bot stopped.")

if __name__ == '__main__':
    if not TELEGRAM_AVAILABLE or not PANDAS_AVAILABLE or not SQLALCHEMY_AVAILABLE:
        sys.exit(1)

    bot = MyanmarFinanceBot()
    try:
        asyncio.run(bot.run())
    except KeyboardInterrupt:
        logger.info("Bot stopped manually from console.")
